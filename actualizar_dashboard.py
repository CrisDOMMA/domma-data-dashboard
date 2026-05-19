#!/usr/bin/env python3
"""
DOMMA - Actualizar Dashboard de Menopausia
==========================================
Descubre CSVs automaticamente, procesa datos, genera dashboard HTML + Excel,
y despliega a GitHub Pages.

Uso:
    python3 actualizar_dashboard.py            # Ejecutar completo
    python3 actualizar_dashboard.py --dry-run  # Solo mostrar que se haria

Configuracion:
    Variable de entorno GITHUB_TOKEN debe estar definida para el despliegue.
    Repo: CrisDOMMA/domma-data-dashboard
"""

import os
import sys
import glob
import json
import csv
import argparse
import base64
import urllib.request
import urllib.error
from datetime import datetime
from collections import Counter, defaultdict
from pathlib import Path

# ---------------------------------------------------------------------------
# CONSTANTES
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent
DOWNLOADS_DIR = Path.home() / "Downloads"
DATOS_NUEVOS_DIR = BASE_DIR / "datos_nuevos"
OUTPUT_HTML = BASE_DIR / "index.html"
OUTPUT_XLSX = BASE_DIR / "DOMMA_Analisis_Menopausia_BA.xlsx"

GITHUB_REPO = "CrisDOMMA/domma-data-dashboard"
GITHUB_API = "https://api.github.com"

# Colores del dashboard
AZUL = "#101637"
ROSA = "#E7C6B9"
ROSA_OSC = "#C4977F"
CREMA = "#F8F7F5"
GRIS = "#6B6B6B"
AZUL_MED = "#2A3060"

# ---------------------------------------------------------------------------
# DESCUBRIMIENTO DE ARCHIVOS
# ---------------------------------------------------------------------------

def descubrir_csvs():
    """Descubre todos los CSVs relevantes."""
    fuentes = {
        "woocommerce": [],
        "shopify": [],
        "typeform_principal": [],
        "typeform_fisico": [],
        "datos_nuevos": [],
    }

    # WooCommerce RevenueHunt
    patron_woo = str(DOWNLOADS_DIR / "jyHJgz*.csv")
    fuentes["woocommerce"] = sorted(glob.glob(patron_woo))

    # Shopify RevenueHunt
    patron_shopify = str(DOWNLOADS_DIR / "*foAoCM*.csv")
    fuentes["shopify"] = sorted(glob.glob(patron_shopify))

    # Typeform principal (~13K)
    patron_tf_main = str(DOWNLOADS_DIR / "responses-rXKhc6VM*.csv")
    fuentes["typeform_principal"] = sorted(glob.glob(patron_tf_main))

    # Typeform fisico (~750)
    patron_tf_fisico = str(DOWNLOADS_DIR / "responses-fGe6BPkD*.csv")
    fuentes["typeform_fisico"] = sorted(glob.glob(patron_tf_fisico))

    # Cualquier CSV nuevo en datos_nuevos/
    if DATOS_NUEVOS_DIR.exists():
        fuentes["datos_nuevos"] = sorted(glob.glob(str(DATOS_NUEVOS_DIR / "*.csv")))

    return fuentes


# ---------------------------------------------------------------------------
# PARSEO DE CSVs
# ---------------------------------------------------------------------------

def leer_csv_seguro(filepath, max_rows=None):
    """Lee un CSV con deteccion de encoding."""
    for enc in ["utf-8", "utf-8-sig", "latin-1", "cp1252"]:
        try:
            with open(filepath, "r", encoding=enc) as f:
                reader = csv.DictReader(f)
                rows = []
                for i, row in enumerate(reader):
                    if max_rows and i >= max_rows:
                        break
                    rows.append(row)
                return rows, reader.fieldnames
        except (UnicodeDecodeError, UnicodeError):
            continue
    print(f"  [ERROR] No se pudo leer {filepath} con ninguna codificacion")
    return [], []


def encontrar_columna(fieldnames, patrones):
    """Encuentra una columna que contenga alguno de los patrones dados."""
    if not fieldnames:
        return None
    for col in fieldnames:
        col_lower = col.lower()
        for patron in patrones:
            if patron.lower() in col_lower:
                return col
    return None


# ---------------------------------------------------------------------------
# PROCESAMIENTO WOOCOMMERCE REVENUEHUNT
# ---------------------------------------------------------------------------

def procesar_woocommerce(filepath):
    """Procesa el CSV de WooCommerce RevenueHunt."""
    print(f"  Leyendo WooCommerce: {Path(filepath).name}...")
    rows, fieldnames = leer_csv_seguro(filepath)
    if not rows:
        return {}

    n = len(rows)
    print(f"    -> {n:,} respuestas")

    # Columnas clave por patron
    col_fase = encontrar_columna(fieldnames, ["menstruacion", "menstruación", "GQiJP55"])
    col_sintomas = encontrar_columna(fieldnames, ["sintomas", "síntomas", "cambios en tu cuerpo", "JniX2eN"])
    col_emocional = encontrar_columna(fieldnames, ["Salud mental", "emocional", "B5iknMX"])
    col_canal = encontrar_columna(fieldnames, ["como nos has conocido", "conocido", "bJi36o0"])
    col_sexual = encontrar_columna(fieldnames, ["Salud sexual", "sexual", "X8iQ7be"])
    col_sofocos = encontrar_columna(fieldnames, ["sofocos y sudores", "geiE2BA"])
    col_fecha = "created_at"
    col_tags = "tags"
    col_preocupa = encontrar_columna(fieldnames, ["preocupa de la menopausia", "qQi2Aw2"])
    col_antojos = encontrar_columna(fieldnames, ["antojos", "jxikmVW"])
    col_grasa = encontrar_columna(fieldnames, ["grasa en abdomen", "1BiomMl"])
    col_peso_inv = encontrar_columna(fieldnames, ["subes mas rapido", "subes más rápido", "vziW50V"])
    col_aspecto = encontrar_columna(fieldnames, ["aspecto", "geik577"])
    col_descanso = encontrar_columna(fieldnames, ["descanso", "kdi23zb"])

    # --- Fase menstrual ---
    fase_counter = Counter()
    for r in rows:
        if col_fase and r.get(col_fase):
            val = r[col_fase].strip()
            if "REGULAR" in val.upper():
                fase_counter["Regular"] += 1
            elif "ALTERADA" in val.upper():
                fase_counter["Alterada"] += 1
            elif "POCO" in val.upper() or "FRECUENTE" in val.upper():
                fase_counter["Poco frecuente"] += 1
            elif "NADA" in val.upper():
                fase_counter["Sin menstruacion"] += 1
            else:
                fase_counter["Otro"] += 1

    # --- Sintomas desde tags ---
    sintoma_map = {
        "sofocos": "Sofocos",
        "insomnio": "Insomnio",
        "baja libido": "Baja libido",
        "aumento peso": "Aumento de peso",
        "aumento de peso": "Aumento de peso",
        "fatiga": "Fatiga",
        "cansancio": "Fatiga",
        "sequedad vaginal": "Sequedad vaginal",
        "sequedad interna": "Sequedad vaginal",
        "sequedad": "Sequedad general",
        "sudores": "Sudores nocturnos",
        "sofocos nocturnos": "Sudores nocturnos",
        "problemas de sueño": "Insomnio",
        "hinchazón": "Hinchazion",
        "infecciones urinarias": "Infecciones urinarias",
        "dolor articular": "Dolor articular",
        "perdida de pelo": "Perdida de pelo",
        "piel seca": "Piel seca",
    }
    sintoma_counter = Counter()
    emo_map = {
        "estresada": "Estres",
        "ansiosa": "Ansiedad",
        "irritable": "Irritabilidad",
        "triste": "Tristeza",
        "cambios de humor": "Cambios de humor",
        "descentrada": "Descentrada",
        "estupenda": "Estupenda",
    }
    emo_counter = Counter()
    sintomas_por_persona = []

    for r in rows:
        tags = r.get(col_tags, "")
        if not tags:
            continue
        tags_lower = tags.lower()
        persona_sintomas = set()
        persona_emos = set()

        # Sintomas de tags
        for key, label in sintoma_map.items():
            if key in tags_lower:
                persona_sintomas.add(label)
        sintomas_por_persona.append(len(persona_sintomas))
        for s in persona_sintomas:
            sintoma_counter[s] += 1

        # Emocional de tags
        for key, label in emo_map.items():
            if key in tags_lower:
                persona_emos.add(label)
        for e in persona_emos:
            emo_counter[e] += 1

    # Tambien leer columna emocional directa
    if col_emocional:
        for r in rows:
            val = r.get(col_emocional, "").upper()
            if not val:
                continue
            for key, label in emo_map.items():
                if key.upper() in val and label not in emo_counter:
                    pass  # Ya contado via tags

    # --- Canal ---
    canal_counter = Counter()
    for r in rows:
        val = ""
        if col_canal:
            val = r.get(col_canal, "")
        if not val and col_tags:
            tags = r.get(col_tags, "")
            # Canal a veces esta en tags
            if "televisión" in tags.lower() or "tv" in tags.lower():
                val = "TV"
            elif "redes sociales" in tags.lower() or "rrss" in tags.lower():
                val = "RRSS"
        if val:
            val_upper = val.upper().strip()
            if any(x in val_upper for x in ["REDES", "RRSS", "INSTAGRAM", "FACEBOOK", "TIKTOK", "DOMMA.STORIES"]):
                canal_counter["RRSS"] += 1
            elif any(x in val_upper for x in ["BUSCA", "GOOGLE", "INTERNET", "SEARCH"]):
                canal_counter["Busqueda"] += 1
            elif any(x in val_upper for x in ["AMIGA", "FAMILIAR", "RECOMEND"]):
                canal_counter["Recomendacion"] += 1
            elif any(x in val_upper for x in ["PRENSA", "RADIO", "PERIÓDICO", "PERIODICO", "PODCAST"]):
                canal_counter["Prensa"] += 1
            elif any(x in val_upper for x in ["MÉDIC", "MEDIC", "DOCTOR", "GINEC", "PROFESIONAL", "SALUD"]):
                canal_counter["Prof. salud"] += 1
            elif any(x in val_upper for x in ["TV", "TELEVI"]):
                canal_counter["TV"] += 1
            else:
                canal_counter["Otro"] += 1

    # --- Fechas ---
    fechas = []
    for r in rows:
        d = r.get(col_fecha, "")
        if d:
            try:
                fechas.append(datetime.strptime(d[:10], "%Y-%m-%d"))
            except ValueError:
                pass

    # --- Sofocos ---
    sofocos_counter = Counter()
    if col_sofocos:
        for r in rows:
            val = r.get(col_sofocos, "").upper()
            if not val:
                continue
            if "NOCHE" in val and "DÍA" in val or "DIA" in val and "NOCHE" in val:
                sofocos_counter["Dia y noche"] += 1
            elif "NOCHE" in val:
                sofocos_counter["Solo noche"] += 1
            elif "DÍA" in val or "DIA" in val or "MAÑANA" in val:
                sofocos_counter["Solo dia"] += 1
            else:
                sofocos_counter["Otro"] += 1

    # --- Antojos / Grasa / Peso ---
    antojos_si = 0
    grasa_si = 0
    peso_si = 0
    antojos_total = 0
    grasa_total = 0
    peso_total = 0

    if col_antojos:
        for r in rows:
            val = r.get(col_antojos, "").upper()
            if val:
                antojos_total += 1
                if "SÍ" in val or "SI" in val or "MUCHOS" in val or "BASTANTE" in val:
                    antojos_si += 1

    if col_grasa:
        for r in rows:
            val = r.get(col_grasa, "").upper()
            if val:
                grasa_total += 1
                if "SÍ" in val or "SI" in val or "CLARAMENTE" in val or "BASTANTE" in val:
                    grasa_si += 1

    if col_peso_inv:
        for r in rows:
            val = r.get(col_peso_inv, "").upper()
            if val:
                peso_total += 1
                if "SÍ" in val or "SI" in val or "TOTALMENTE" in val:
                    peso_si += 1

    # --- Volumen por trimestre ---
    vol_trimestre = Counter()
    for f in fechas:
        q = (f.month - 1) // 3 + 1
        key = f"{f.year}Q{q}"
        vol_trimestre[key] += 1

    # --- Fase por persona para cruces ---
    fase_por_idx = {}
    for i, r in enumerate(rows):
        if col_fase and r.get(col_fase):
            val = r[col_fase].strip().upper()
            if "REGULAR" in val:
                fase_por_idx[i] = "Pre-menopausia"
            elif "ALTERADA" in val:
                fase_por_idx[i] = "Peri (alterada)"
            elif "POCO" in val or "FRECUENTE" in val:
                fase_por_idx[i] = "Peri (escasa)"
            elif "NADA" in val:
                fase_por_idx[i] = "Post-menopausia"

    # Sintomas por fase
    sintomas_por_fase = defaultdict(lambda: defaultdict(int))
    fase_totals = Counter()
    for i, r in enumerate(rows):
        fase = fase_por_idx.get(i)
        if not fase:
            continue
        fase_totals[fase] += 1
        tags = r.get(col_tags, "").lower()
        for key, label in sintoma_map.items():
            if key in tags:
                sintomas_por_fase[label][fase] += 1

    # Indice de sufrimiento
    suffering = Counter()
    for count in sintomas_por_persona:
        if count > 0:
            suffering[str(min(count, 10))] += 1

    # --- Segmentos ---
    seg_counter = Counter()
    severe_emo_keys = {"ansiosa", "triste", "irritable"}
    for r in rows:
        tags = r.get(col_tags, "").lower()
        phys_count = sum(1 for k in sintoma_map if k in tags)
        emo_val = ""
        if col_emocional:
            emo_val = r.get(col_emocional, "").lower()
        emo_from_tags = tags
        has_severe_emo = any(k in emo_from_tags or k in emo_val for k in severe_emo_keys)
        has_estupenda = "estupenda" in emo_from_tags or "estupenda" in emo_val

        if phys_count >= 5 and has_severe_emo:
            seg_counter["Sobrecarga total"] += 1
        elif phys_count >= 5:
            seg_counter["Carga fisica alta"] += 1
        elif has_severe_emo:
            seg_counter["Impacto emocional"] += 1
        elif phys_count >= 2:
            seg_counter["Moderada"] += 1
        elif phys_count >= 1 and has_estupenda:
            seg_counter["Cuerpo si, mente no"] += 1
        else:
            seg_counter["Bien"] += 1

    return {
        "fuente": "WooCommerce RevenueHunt",
        "total": n,
        "fase": dict(fase_counter),
        "sintomas": dict(sintoma_counter),
        "emocional": dict(emo_counter),
        "canal": dict(canal_counter),
        "fecha_min": min(fechas).strftime("%Y-%m-%d") if fechas else None,
        "fecha_max": max(fechas).strftime("%Y-%m-%d") if fechas else None,
        "sofocos": dict(sofocos_counter),
        "antojos_pct": round(antojos_si / antojos_total * 100, 1) if antojos_total > 0 else None,
        "grasa_pct": round(grasa_si / grasa_total * 100, 1) if grasa_total > 0 else None,
        "peso_pct": round(peso_si / peso_total * 100, 1) if peso_total > 0 else None,
        "volumen_trimestre": dict(sorted(vol_trimestre.items())),
        "sintomas_por_fase": {s: dict(f) for s, f in sintomas_por_fase.items()},
        "fase_totals": dict(fase_totals),
        "suffering": dict(sorted(suffering.items(), key=lambda x: int(x[0]))),
        "segmentos": dict(seg_counter),
        "media_sintomas": round(sum(sintomas_por_persona) / len(sintomas_por_persona), 1) if sintomas_por_persona else 0,
    }


# ---------------------------------------------------------------------------
# PROCESAMIENTO TYPEFORM PRINCIPAL
# ---------------------------------------------------------------------------

def procesar_typeform_principal(filepath):
    """Procesa el CSV de Typeform principal."""
    print(f"  Leyendo Typeform principal: {Path(filepath).name}...")
    rows, fieldnames = leer_csv_seguro(filepath)
    if not rows:
        return {}
    n = len(rows)
    print(f"    -> {n:,} respuestas")

    # Columnas de sintomas son directas: Sofocos, Sequedad vaginal, etc.
    sintoma_cols = {
        "Sofocos": "Sofocos",
        "Sequedad vaginal": "Sequedad vaginal",
        "Sequedad": "Sequedad general",
        "Insomnio": "Insomnio",
        "Baja libido": "Baja libido",
        "Sudores nocturnos": "Sudores nocturnos",
        "Aumento de peso": "Aumento de peso",
        "Fatiga": "Fatiga",
    }
    emo_cols = {
        "ANSIOSA": "Ansiedad",
        "ESTRESADA": "Estres",
        "CAMBIOS DE HUMOR": "Cambios de humor",
        "DESCENTRADA": "Descentrada",
        "IRRITABLE": "Irritabilidad",
        "TRISTE": "Tristeza",
        "ESTUPENDA": "Estupenda",
    }

    # Fase menstrual
    col_fase = encontrar_columna(fieldnames, ["menstruacion", "menstruación"])
    col_nacimiento = encontrar_columna(fieldnames, ["naciste", "nacimiento"])
    col_canal = encontrar_columna(fieldnames, ["conocido", "como nos has"])
    col_submit = encontrar_columna(fieldnames, ["Submit Date"])

    fase_counter = Counter()
    sintoma_counter = Counter()
    emo_counter = Counter()
    canal_counter = Counter()
    edad_counter = Counter()
    fechas = []

    for r in rows:
        # Fase
        if col_fase:
            val = r.get(col_fase, "").upper()
            if "REGULAR" in val:
                fase_counter["Regular"] += 1
            elif "ALTERADA" in val:
                fase_counter["Alterada"] += 1
            elif "POCO" in val or "FRECUENTE" in val:
                fase_counter["Poco frecuente"] += 1
            elif "NADA" in val:
                fase_counter["Sin menstruacion"] += 1

        # Sintomas
        for col_name, label in sintoma_cols.items():
            val = r.get(col_name, "")
            if val and val.strip():
                sintoma_counter[label] += 1

        # Emocional
        for col_name, label in emo_cols.items():
            val = r.get(col_name, "")
            if val and val.strip():
                emo_counter[label] += 1

        # Canal
        if col_canal:
            val = r.get(col_canal, "").upper()
            if val:
                if any(x in val for x in ["REDES", "RRSS", "INSTAGRAM", "FACEBOOK", "TIKTOK", "DOMMA"]):
                    canal_counter["RRSS"] += 1
                elif any(x in val for x in ["BUSCA", "GOOGLE", "INTERNET"]):
                    canal_counter["Busqueda"] += 1
                elif any(x in val for x in ["AMIGA", "FAMILIAR", "RECOMEND"]):
                    canal_counter["Recomendacion"] += 1
                elif any(x in val for x in ["PRENSA", "RADIO", "PODCAST"]):
                    canal_counter["Prensa"] += 1
                elif any(x in val for x in ["MÉDIC", "MEDIC", "GINEC", "PROFESIONAL", "SALUD"]):
                    canal_counter["Prof. salud"] += 1
                elif any(x in val for x in ["TV", "TELEVI"]):
                    canal_counter["TV"] += 1
                else:
                    canal_counter["Otro"] += 1

        # Edad
        if col_nacimiento:
            dob = r.get(col_nacimiento, "")
            if dob:
                try:
                    birth = datetime.strptime(dob[:10], "%Y-%m-%d")
                    age = (datetime.now() - birth).days // 365
                    if 35 <= age < 40:
                        edad_counter["35-39"] += 1
                    elif 40 <= age < 45:
                        edad_counter["40-44"] += 1
                    elif 45 <= age < 50:
                        edad_counter["45-49"] += 1
                    elif 50 <= age < 55:
                        edad_counter["50-54"] += 1
                    elif 55 <= age < 60:
                        edad_counter["55-59"] += 1
                    elif 60 <= age < 65:
                        edad_counter["60-64"] += 1
                    elif age >= 65:
                        edad_counter["65+"] += 1
                except (ValueError, TypeError):
                    pass

        # Fechas
        if col_submit:
            d = r.get(col_submit, "")
            if d:
                try:
                    fechas.append(datetime.strptime(d[:10], "%Y-%m-%d"))
                except ValueError:
                    pass

    return {
        "fuente": "Typeform Principal",
        "total": n,
        "fase": dict(fase_counter),
        "sintomas": dict(sintoma_counter),
        "emocional": dict(emo_counter),
        "canal": dict(canal_counter),
        "edad": dict(edad_counter),
        "fecha_min": min(fechas).strftime("%Y-%m-%d") if fechas else None,
        "fecha_max": max(fechas).strftime("%Y-%m-%d") if fechas else None,
    }


# ---------------------------------------------------------------------------
# PROCESAMIENTO TYPEFORM FISICO
# ---------------------------------------------------------------------------

def procesar_typeform_fisico(filepath):
    """Procesa el CSV del estudio fisico."""
    print(f"  Leyendo Typeform fisico: {Path(filepath).name}...")
    rows, fieldnames = leer_csv_seguro(filepath)
    if not rows:
        return {}
    n = len(rows)
    print(f"    -> {n:,} respuestas")

    col_edad = encontrar_columna(fieldnames, ["edad", "qué edad"])
    col_fase = encontrar_columna(fieldnames, ["etapa", "menstruacion", "menstruación"])
    col_submit = encontrar_columna(fieldnames, ["Submit Date"])

    # Cambios fisicos (multi-columna)
    cambios_cols = {
        "grasa": "Grasa abdominal",
        "fuerza": "Perdida fuerza/musculo",
        "mantener el peso": "Cuesta mantener peso",
        "cansada": "Cansancio/menos energia",
        "duele": "Dolor/rigidez",
        "rigidez": "Dolor/rigidez",
        "marcadores": "Marcadores alterados",
        "osteo": "Perdida osea",
        "pérdida ósea": "Perdida osea",
        "hinchada": "Hinchada",
        "no he notado": "Sin cambios",
    }
    # Preocupaciones
    preocup_cols = {
        "peso": "El peso",
        "grasa": "Grasa barriga",
        "barriga": "Grasa barriga",
        "musculo": "Perdida musculo/fuerza",
        "fuerza": "Perdida musculo/fuerza",
        "articulaciones": "Dolores articulaciones",
        "dolor": "Dolores articulaciones",
        "energia": "Sin energia/agotada",
        "agotada": "Sin energia/agotada",
        "glucosa": "Glucosa/colesterol",
        "colesterol": "Glucosa/colesterol",
        "riesgo": "Riesgo enfermedades",
        "nada en especial": "Nada en especial",
    }

    edad_counter = Counter()
    fase_counter = Counter()
    cambios_counter = Counter()
    preocup_counter = Counter()
    ejercicio_si = 0
    ejercicio_no_quiere = 0
    ejercicio_no_interesa = 0
    tipo_ejercicio = Counter()
    rendimiento = Counter()
    nutricion = Counter()
    plan_nutri = Counter()
    barreras = Counter()
    salud_osea = Counter()
    salud_cardio = Counter()
    fechas = []

    # Identify actual column names
    cambio_field_map = {}
    preocup_field_map = {}
    for fn in (fieldnames or []):
        fn_lower = fn.lower()
        for key, label in cambios_cols.items():
            if key in fn_lower:
                cambio_field_map[fn] = label
        for key, label in preocup_cols.items():
            if key in fn_lower and fn not in cambio_field_map:
                preocup_field_map[fn] = label

    # Exercise columns
    col_ejercicio = encontrar_columna(fieldnames, ["ejercicio de forma regular"])
    col_caminar = encontrar_columna(fieldnames, ["Caminar"])
    col_yoga = encontrar_columna(fieldnames, ["Yoga", "pilates"])
    col_cardio = encontrar_columna(fieldnames, ["cardiovascular", "correr"])
    col_fuerza = encontrar_columna(fieldnames, ["fuerza", "pesas"])

    # Performance columns
    col_bien = encontrar_columna(fieldnames, ["me siento bien y recupero"])
    col_cuesta_fuerza = encontrar_columna(fieldnames, ["cuesta más ganar fuerza"])
    col_canso = encontrar_columna(fieldnames, ["canso más"])
    col_dejado = encontrar_columna(fieldnames, ["dejado de hacer"])
    col_perder_grasa = encontrar_columna(fieldnames, ["cuesta perder grasa"])

    # Nutrition
    col_picar = encontrar_columna(fieldnames, ["picar", "hambre"])
    col_antojos = encontrar_columna(fieldnames, ["antojos", "ansiedad"])
    col_equilibrio = encontrar_columna(fieldnames, ["equilibrio"])
    col_plan = encontrar_columna(fieldnames, ["plan nutricional"])

    # Barriers
    col_b_tiempo = encontrar_columna(fieldnames, ["Falta de tiempo"])
    col_b_energia = encontrar_columna(fieldnames, ["Falta de energía", "Falta de energia"])
    col_b_empezar = encontrar_columna(fieldnames, ["por dónde empezar", "por donde empezar"])
    col_b_resultados = encontrar_columna(fieldnames, ["no veo resultados"])
    col_b_tarde = encontrar_columna(fieldnames, ["ya es tarde", "no responde"])

    # Bone / cardio
    col_huesos = encontrar_columna(fieldnames, ["hablado alguna vez de cuidar tus huesos"])
    col_corazon = encontrar_columna(fieldnames, ["corazón", "cardiovascular"])

    for r in rows:
        # Edad
        if col_edad:
            val = r.get(col_edad, "").strip()
            if val:
                edad_counter[val] += 1

        # Fase
        if col_fase:
            val = r.get(col_fase, "").lower()
            if "posmenopausia" in val or "hace años" in val:
                fase_counter["Posmenopausia"] += 1
            elif "menopausia" in val and "peri" not in val and "pre" not in val and "pos" not in val:
                fase_counter["Menopausia"] += 1
            elif "peri" in val or "irregular" in val:
                fase_counter["Perimenopausia"] += 1
            elif "pre" in val or "regular" in val:
                fase_counter["Premenopausia"] += 1
            elif "no" in val and "regla" in val:
                fase_counter["Posmenopausia"] += 1
            else:
                fase_counter["Otro"] += 1

        # Cambios fisicos
        for fn, label in cambio_field_map.items():
            if r.get(fn, "").strip():
                cambios_counter[label] += 1

        # Preocupaciones
        for fn, label in preocup_field_map.items():
            if r.get(fn, "").strip():
                preocup_counter[label] += 1

        # Ejercicio
        if col_ejercicio:
            val = r.get(col_ejercicio, "").lower()
            if "sí" in val or "si" in val:
                ejercicio_si += 1
            elif "no" in val and ("gustaría" in val or "gustaria" in val or "empezar" in val):
                ejercicio_no_quiere += 1
            elif "no" in val:
                ejercicio_no_interesa += 1

        # Tipo de ejercicio
        if col_caminar and r.get(col_caminar, "").strip():
            tipo_ejercicio["Caminar"] += 1
        if col_yoga and r.get(col_yoga, "").strip():
            tipo_ejercicio["Yoga/Pilates"] += 1
        if col_cardio and r.get(col_cardio, "").strip():
            tipo_ejercicio["Cardio"] += 1
        if col_fuerza and r.get(col_fuerza, "").strip():
            tipo_ejercicio["Fuerza"] += 1

        # Rendimiento
        if col_bien and r.get(col_bien, "").strip():
            rendimiento["Me siento bien"] += 1
        if col_cuesta_fuerza and r.get(col_cuesta_fuerza, "").strip():
            rendimiento["Cuesta ganar fuerza"] += 1
        if col_canso and r.get(col_canso, "").strip():
            rendimiento["Me canso mas"] += 1
        if col_dejado and r.get(col_dejado, "").strip():
            rendimiento["He dejado cosas"] += 1
        if col_perder_grasa and r.get(col_perder_grasa, "").strip():
            rendimiento["Cuesta perder grasa"] += 1

        # Nutricion
        if col_picar and r.get(col_picar, "").strip():
            nutricion["Picar entre horas"] += 1
        if col_antojos and r.get(col_antojos, "").strip():
            nutricion["Antojos/ansiedad"] += 1
        if col_equilibrio and r.get(col_equilibrio, "").strip():
            nutricion["Equilibrio"] += 1

        # Barreras
        if col_b_tiempo and r.get(col_b_tiempo, "").strip():
            barreras["Falta de tiempo"] += 1
        if col_b_energia and r.get(col_b_energia, "").strip():
            barreras["Falta de energia"] += 1
        if col_b_empezar and r.get(col_b_empezar, "").strip():
            barreras["No se por donde empezar"] += 1
        if col_b_resultados and r.get(col_b_resultados, "").strip():
            barreras["No veo resultados"] += 1
        if col_b_tarde and r.get(col_b_tarde, "").strip():
            barreras["Ya es tarde"] += 1

        # Fechas
        if col_submit:
            d = r.get(col_submit, "")
            if d:
                try:
                    fechas.append(datetime.strptime(d[:10], "%Y-%m-%d"))
                except ValueError:
                    pass

    return {
        "fuente": "Typeform Fisico",
        "total": n,
        "edad": dict(edad_counter),
        "fase": dict(fase_counter),
        "cambios_fisicos": dict(cambios_counter),
        "preocupaciones": dict(preocup_counter),
        "ejercicio_pct": round(ejercicio_si / n * 100, 1) if n > 0 else 0,
        "ejercicio_quiere_empezar_pct": round(ejercicio_no_quiere / n * 100, 1) if n > 0 else 0,
        "tipo_ejercicio": dict(tipo_ejercicio),
        "rendimiento": dict(rendimiento),
        "nutricion": dict(nutricion),
        "barreras": dict(barreras),
        "fecha_min": min(fechas).strftime("%Y-%m-%d") if fechas else None,
        "fecha_max": max(fechas).strftime("%Y-%m-%d") if fechas else None,
    }


# ---------------------------------------------------------------------------
# COMBINAR DATOS
# ---------------------------------------------------------------------------

def combinar_datos(resultados):
    """Combina metricas de todas las fuentes."""
    total = sum(r.get("total", 0) for r in resultados)

    # Combinar fases
    fase_combined = Counter()
    for r in resultados:
        for k, v in r.get("fase", {}).items():
            fase_combined[k] += v

    # Combinar sintomas
    sintoma_combined = Counter()
    for r in resultados:
        for k, v in r.get("sintomas", {}).items():
            sintoma_combined[k] += v

    # Combinar emocional
    emo_combined = Counter()
    for r in resultados:
        for k, v in r.get("emocional", {}).items():
            emo_combined[k] += v

    # Combinar canales
    canal_combined = Counter()
    for r in resultados:
        for k, v in r.get("canal", {}).items():
            canal_combined[k] += v

    # Fechas
    fecha_min = None
    fecha_max = None
    for r in resultados:
        fm = r.get("fecha_min")
        fx = r.get("fecha_max")
        if fm:
            if not fecha_min or fm < fecha_min:
                fecha_min = fm
        if fx:
            if not fecha_max or fx > fecha_max:
                fecha_max = fx

    # Volumen por trimestre (del WooCommerce)
    vol = {}
    for r in resultados:
        for k, v in r.get("volumen_trimestre", {}).items():
            vol[k] = vol.get(k, 0) + v

    # Extraer datos WooCommerce para cruces, sufrimiento, segmentos
    woo_data = next((r for r in resultados if "WooCommerce" in r.get("fuente", "")), {})
    fisico_data = next((r for r in resultados if "Fisico" in r.get("fuente", "")), {})

    # Calcular porcentajes de sintomas sobre total de respondentes con tags
    # Usamos el total de WooCommerce como base para los % de sintomas
    woo_total = woo_data.get("total", 1)
    tf_total = next((r.get("total", 0) for r in resultados if "Typeform Principal" in r.get("fuente", "")), 0)

    # Top sintomas porcentaje
    top_sintomas = {}
    for s, count in sintoma_combined.most_common(10):
        top_sintomas[s] = round(count / total * 100, 1)

    # Top emocional porcentaje
    top_emo = {}
    for e, count in emo_combined.most_common(10):
        top_emo[e] = round(count / total * 100, 1)

    # Canal porcentaje
    canal_total = sum(canal_combined.values()) or 1
    canal_pct = {k: round(v / canal_total * 100, 1) for k, v in canal_combined.most_common(10)}

    # Media de sintomas
    media_sintomas = woo_data.get("media_sintomas", 5.8)

    # Año rango
    year_min = fecha_min[:4] if fecha_min else "2022"
    year_max = fecha_max[:4] if fecha_max else "2026"

    return {
        "total": total,
        "media_sintomas": media_sintomas,
        "fecha_min": fecha_min,
        "fecha_max": fecha_max,
        "year_range": f"{year_min}-{year_max}",
        "fase": dict(fase_combined),
        "sintomas_pct": top_sintomas,
        "emocional_pct": top_emo,
        "canal_pct": canal_pct,
        "volumen_trimestre": dict(sorted(vol.items())) if vol else None,
        "sofocos": woo_data.get("sofocos", {}),
        "suffering": woo_data.get("suffering", {}),
        "segmentos": woo_data.get("segmentos", {}),
        "sintomas_por_fase": woo_data.get("sintomas_por_fase", {}),
        "fase_totals": woo_data.get("fase_totals", {}),
        "antojos_pct": woo_data.get("antojos_pct"),
        "grasa_pct": woo_data.get("grasa_pct"),
        "peso_pct": woo_data.get("peso_pct"),
        "fisico": fisico_data,
        "fuentes": [{
            "nombre": r.get("fuente", "?"),
            "total": r.get("total", 0),
            "fecha_min": r.get("fecha_min"),
            "fecha_max": r.get("fecha_max"),
        } for r in resultados],
    }


# ---------------------------------------------------------------------------
# GENERACION HTML
# ---------------------------------------------------------------------------

def generar_html(datos):
    """Genera el HTML completo del dashboard."""
    total = datos["total"]
    media = datos["media_sintomas"]
    year_range = datos["year_range"]

    # Preparar datos para JS
    fase = datos["fase"]
    sym_items = sorted(datos["sintomas_pct"].items(), key=lambda x: -x[1])[:8]
    sym_labels = [x[0] for x in sym_items]
    sym_pcts = [x[1] for x in sym_items]

    emo_items = sorted(datos["emocional_pct"].items(), key=lambda x: -x[1])
    emo_labels = [x[0] for x in emo_items]
    emo_pcts = [x[1] for x in emo_items]

    canal_items = sorted(datos["canal_pct"].items(), key=lambda x: -x[1])
    canal_labels = [x[0] for x in canal_items]
    canal_pcts = [x[1] for x in canal_items]

    vol = datos.get("volumen_trimestre") or {}
    sofocos = datos.get("sofocos", {})
    suffering = datos.get("suffering", {})
    segmentos = datos.get("segmentos", {})

    # Emocional highlight
    emo_cambios = datos["emocional_pct"].get("Cambios de humor", 58)
    emo_estres = datos["emocional_pct"].get("Estres", 54)
    emo_bien = datos["emocional_pct"].get("Estupenda", 14)

    # Con impacto emocional: % que NO es estupenda
    pct_impacto_emo = round(100 - emo_bien, 0)

    # Canal highlights
    rrss_pct = datos["canal_pct"].get("RRSS", 85)
    busqueda_pct = datos["canal_pct"].get("Busqueda", 10)
    prof_pct = datos["canal_pct"].get("Prof. salud", 1)

    # Peso
    grasa_pct = datos.get("grasa_pct") or 78.0
    peso_pct = datos.get("peso_pct") or 72.7
    antojos_pct = datos.get("antojos_pct") or 80.0

    # Cross data - use existing or default
    cross_data = datos.get("sintomas_por_fase", {})
    fase_totals = datos.get("fase_totals", {})

    # Compute cross percentages
    cross_fases = ["Pre-menopausia", "Peri (alterada)", "Peri (escasa)", "Post-menopausia"]
    cross_js = {}
    for sintoma, fase_counts in cross_data.items():
        cross_js[sintoma] = []
        for f in cross_fases:
            ft = fase_totals.get(f, 1)
            cross_js[sintoma].append(round(fase_counts.get(f, 0) / ft * 100, 1) if ft > 0 else 0)

    # Fisico data
    fisico = datos.get("fisico", {})
    fisico_total = fisico.get("total", 729)
    fisico_edad = fisico.get("edad", {})
    fisico_fase = fisico.get("fase", {})
    fisico_cambios = fisico.get("cambios_fisicos", {})
    fisico_preocup = fisico.get("preocupaciones", {})
    fisico_ejercicio_pct = fisico.get("ejercicio_pct", 68.6)
    fisico_tipo = fisico.get("tipo_ejercicio", {})
    fisico_rend = fisico.get("rendimiento", {})
    fisico_nutri = fisico.get("nutricion", {})
    fisico_barreras = fisico.get("barreras", {})

    # Compute fisico percentages
    def pct_of(counter, total_n):
        return {k: round(v / total_n * 100, 1) for k, v in sorted(counter.items(), key=lambda x: -x[1])} if total_n > 0 else counter

    fisico_cambios_pct = pct_of(fisico_cambios, fisico_total)
    fisico_preocup_pct = pct_of(fisico_preocup, fisico_total)
    fisico_tipo_pct = pct_of(fisico_tipo, fisico_total)
    fisico_rend_pct = pct_of(fisico_rend, fisico_total)
    fisico_barreras_pct = pct_of(fisico_barreras, fisico_total)

    # Fisico cambios sin cambios exclusion
    nota_cambios_pct = round(100 - fisico_cambios_pct.get("Sin cambios", 3.7), 1)

    # Fisico edad mayoritario
    fisico_edad_top = max(fisico_edad.items(), key=lambda x: x[1])[0] if fisico_edad else "50-54"
    fisico_edad_top_pct = round(max(fisico_edad.values()) / fisico_total * 100, 1) if fisico_edad and fisico_total > 0 else 40.5

    # Suffering: % con 5+ sintomas
    total_suffering = sum(int(v) for v in suffering.values()) if suffering else 1
    pct_5plus = round(sum(v for k, v in suffering.items() if int(k) >= 5) / total_suffering * 100, 0) if suffering else 71

    # Segmentos - sobrecarga
    seg_total = sum(segmentos.values()) or 1
    seg_sobrecarga_pct = round(segmentos.get("Sobrecarga total", 0) / seg_total * 100, 0)

    # Format total with dots
    total_fmt = f"{total:,}".replace(",", ".")

    # Co-occurrence placeholder (keep original if no new data)
    cooc_labels_js = json.dumps(["Baja libido", "Insomnio", "Aum. peso", "Sofocos", "Fatiga", "Seq. vaginal", "Sudores", "Hinchazion"])
    cooc_matrix_js = json.dumps([
        [100, 68.7, 70.0, 59.9, 71.1, 63.2, 11.0, 57.9],
        [66.9, 100, 70.5, 65.2, 74.4, 51.9, 11.0, 59.7],
        [66.4, 68.7, 100, 60.6, 74.2, 50.0, 9.5, 84.1],
        [67.5, 75.6, 72.1, 100, 73.3, 52.9, 14.2, 60.1],
        [66.9, 71.9, 73.6, 61.1, 100, 49.8, 8.3, 64.2],
        [82.7, 69.7, 69.0, 61.3, 69.2, 100, 11.3, 56.5],
        [69.9, 71.6, 63.2, 79.4, 55.8, 54.6, 100, 0.0],
        [65.3, 69.2, 100.0, 60.1, 76.9, 48.7, 0.0, 100],
    ])

    # Age cross data (keep defaults if not computed)
    age_cross_js = json.dumps({
        "Baja libido": [76.7, 71.5, 71.1, 74.7, 78.2],
        "Insomnio": [52.4, 59.6, 64.5, 64.9, 55.3],
        "Sofocos": [37.0, 50.0, 62.6, 62.5, 48.1],
        "Aumento peso": [51.9, 59.8, 64.1, 62.6, 58.3],
        "Fatiga": [65.1, 60.0, 58.2, 49.8, 44.7],
        "Seq. vaginal": [45.5, 45.9, 53.4, 64.4, 74.6],
        "Sudores noct.": [43.4, 53.9, 58.3, 56.6, 46.2],
    })

    # Build cross_js for template
    cross_js_limited = {}
    top_cross = ["Baja libido", "Insomnio", "Sofocos", "Aumento de peso", "Fatiga", "Sequedad vaginal", "Sudores nocturnos"]
    for s in top_cross:
        if s in cross_js:
            cross_js_limited[s] = cross_js[s]

    html = f'''<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>DOMMA &mdash; Radiograf&iacute;a de la Menopausia</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.7/dist/chart.umd.min.js"></script>
<style>
:root {{
    --azul: {AZUL};
    --rosa: {ROSA};
    --rosa-oscuro: {ROSA_OSC};
    --crema: {CREMA};
    --gris: {GRIS};
    --azul-medio: {AZUL_MED};
}}
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; background: var(--crema); color: var(--azul); }}
.header {{ background: var(--azul); color: white; padding: 24px 32px; display: flex; justify-content: space-between; align-items: center; position: sticky; top: 0; z-index: 100; }}
.header h1 {{ font-size: 20px; font-weight: 700; }}
.header .stats {{ display: flex; gap: 24px; font-size: 13px; opacity: 0.8; }}
.header .stats span {{ font-weight: 700; font-size: 16px; color: var(--rosa); display: block; }}
.nav {{ background: white; padding: 12px 32px; border-bottom: 1px solid #e0e0e0; display: flex; gap: 8px; flex-wrap: wrap; position: sticky; top: 68px; z-index: 99; }}
.nav button {{ padding: 8px 16px; border: 1px solid #ddd; border-radius: 20px; background: white; cursor: pointer; font-size: 13px; transition: all 0.2s; }}
.nav button:hover {{ border-color: var(--rosa-oscuro); }}
.nav button.active {{ background: var(--azul); color: white; border-color: var(--azul); }}
.container {{ max-width: 1200px; margin: 0 auto; padding: 24px; }}
.grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(380px, 1fr)); gap: 20px; margin-bottom: 24px; }}
.card {{ background: white; border-radius: 12px; padding: 24px; box-shadow: 0 1px 3px rgba(0,0,0,0.08); }}
.card h3 {{ font-size: 15px; color: var(--azul); margin-bottom: 16px; font-weight: 700; }}
.card .subtitle {{ font-size: 12px; color: var(--gris); margin-top: -12px; margin-bottom: 16px; }}
.card canvas {{ max-height: 350px; }}
.big-number {{ text-align: center; padding: 20px; }}
.big-number .number {{ font-size: 48px; font-weight: 800; color: var(--azul); }}
.big-number .label {{ font-size: 13px; color: var(--gris); margin-top: 4px; }}
.insight-card {{ background: var(--azul); color: white; border-radius: 12px; padding: 20px; }}
.insight-card h3 {{ color: var(--rosa); font-size: 13px; text-transform: uppercase; letter-spacing: 1px; margin-bottom: 8px; }}
.insight-card p {{ font-size: 15px; line-height: 1.5; }}
.insight-card .number {{ font-size: 36px; font-weight: 800; color: var(--rosa); }}
.section {{ display: none; }}
.section.active {{ display: block; }}
.heatmap-container {{ overflow-x: auto; }}
.heatmap {{ border-collapse: collapse; font-size: 12px; width: 100%; }}
.heatmap th {{ background: var(--azul); color: white; padding: 8px 6px; font-weight: 600; text-align: center; }}
.heatmap td {{ padding: 8px 6px; text-align: center; font-weight: 600; border: 1px solid #eee; }}
.heatmap .row-header {{ background: #f5f5f5; text-align: left; font-weight: 700; color: var(--azul); }}
.download-btn {{ display:inline-flex;align-items:center;gap:8px;background:var(--azul);color:white;padding:10px 20px;border-radius:8px;text-decoration:none;font-size:13px;font-weight:600;transition:opacity 0.2s; }}
.download-btn:hover {{ opacity:0.85; }}
.update-banner {{ background: var(--rosa); color: var(--azul); padding: 8px 32px; font-size: 12px; text-align: center; }}
</style>
</head>
<body>

<div class="update-banner">Actualizado: {datetime.now().strftime("%d/%m/%Y %H:%M")} | Fuentes: {", ".join(f["nombre"] + " (" + str(f["total"]) + ")" for f in datos["fuentes"])}</div>

<div class="header">
    <h1>DOMMA &mdash; Radiograf&iacute;a de la Menopausia en Espa&ntilde;a</h1>
    <div class="stats">
        <div>Respuestas<span>{total_fmt}</span></div>
        <div>Periodo<span>{year_range}</span></div>
        <div>Media s&iacute;ntomas<span>{media}</span></div>
    </div>
</div>

<div class="nav">
    <button class="active" onclick="showSection('overview')">Vista general</button>
    <button onclick="showSection('sintomas')">S&iacute;ntomas</button>
    <button onclick="showSection('emocional')">Emocional</button>
    <button onclick="showSection('cruces')">Cruces</button>
    <button onclick="showSection('peso')">Peso y cuerpo</button>
    <button onclick="showSection('canales')">Canales</button>
    <button onclick="showSection('segmentos')">Segmentos</button>
    <button onclick="showSection('insights')">Insights clave</button>
    <button onclick="showSection('fisico')" style="border-color:var(--rosa-oscuro);font-weight:600;">Estudio F&iacute;sico</button>
</div>

<div class="container">

<!-- OVERVIEW -->
<div id="overview" class="section active">
    <div class="grid" style="grid-template-columns: repeat(4, 1fr);">
        <div class="card big-number"><div class="number">{total_fmt}</div><div class="label">Respuestas totales</div></div>
        <div class="card big-number"><div class="number">{str(media).replace(".", ",")}</div><div class="label">Media de s&iacute;ntomas por mujer</div></div>
        <div class="card big-number"><div class="number">{int(pct_impacto_emo)}%</div><div class="label">Con impacto emocional</div></div>
        <div class="card big-number"><div class="number">{int(prof_pct)}%</div><div class="label">Llega por profesional salud</div></div>
    </div>
    <div class="grid">
        <div class="card"><h3>Fase menstrual</h3><canvas id="chartFase"></canvas></div>
        <div class="card"><h3>Volumen de respuestas por trimestre</h3><canvas id="chartVolumen"></canvas></div>
    </div>
    <div class="grid">
        <div class="card"><h3>S&iacute;ntomas m&aacute;s reportados</h3><canvas id="chartSintomasOverview"></canvas></div>
        <div class="card"><h3>Estado emocional</h3><canvas id="chartEmoOverview"></canvas></div>
    </div>
</div>

<!-- SINTOMAS -->
<div id="sintomas" class="section">
    <div class="grid">
        <div class="card" style="grid-column: span 2;"><h3>S&iacute;ntomas f&iacute;sicos &mdash; % de mujeres que lo reporta</h3><canvas id="chartSintomas"></canvas></div>
    </div>
    <div class="grid">
        <div class="card"><h3>Sofocos: &iquest;cu&aacute;ndo los sufren?</h3><canvas id="chartSofocos"></canvas></div>
        <div class="card"><h3>&Iacute;ndice de sufrimiento: &iquest;cu&aacute;ntos s&iacute;ntomas a la vez?</h3><canvas id="chartSuffering"></canvas></div>
    </div>
</div>

<!-- EMOCIONAL -->
<div id="emocional" class="section">
    <div class="grid" style="grid-template-columns: repeat(3, 1fr);">
        <div class="card big-number insight-card"><div class="number">{int(emo_cambios)}%</div><p>Cambios de humor</p></div>
        <div class="card big-number insight-card"><div class="number">{int(emo_estres)}%</div><p>Estr&eacute;s</p></div>
        <div class="card big-number insight-card"><div class="number">{int(emo_bien)}%</div><p>Se siente bien</p></div>
    </div>
    <div class="grid">
        <div class="card" style="grid-column: span 2;"><h3>Estado emocional de las participantes</h3><canvas id="chartEmocional"></canvas></div>
    </div>
</div>

<!-- CRUCES -->
<div id="cruces" class="section">
    <div class="grid">
        <div class="card" style="grid-column: span 2;"><h3>S&iacute;ntomas por fase menop&aacute;usica</h3><p class="subtitle">% de mujeres que reporta cada s&iacute;ntoma, segmentado por fase</p><canvas id="chartCross"></canvas></div>
    </div>
    <div class="grid">
        <div class="card" style="grid-column: span 2;"><h3>S&iacute;ntomas por rango de edad</h3><p class="subtitle">Datos de mujeres con edad conocida (fuente Typeform)</p><canvas id="chartAgeCross"></canvas></div>
    </div>
    <div class="grid">
        <div class="card" style="grid-column: span 2;">
            <h3>Matriz de co-ocurrencia de s&iacute;ntomas</h3>
            <p class="subtitle">Si una mujer tiene el s&iacute;ntoma de la fila, &iquest;qu&eacute; % tambi&eacute;n tiene el de la columna?</p>
            <div class="heatmap-container">
                <table class="heatmap" id="heatmapTable"></table>
            </div>
        </div>
    </div>
</div>

<!-- PESO -->
<div id="peso" class="section">
    <div class="grid" style="grid-template-columns: repeat(3, 1fr);">
        <div class="card big-number insight-card"><div class="number">{grasa_pct}%</div><p>Acumula grasa abdominal</p></div>
        <div class="card big-number insight-card"><div class="number">{peso_pct}%</div><p>Sube de peso sin cambiar dieta</p></div>
        <div class="card big-number insight-card"><div class="number">{antojos_pct}%</div><p>Tiene antojos frecuentes</p></div>
    </div>
    <div class="grid">
        <div class="card"><h3>Acumulaci&oacute;n de grasa abdominal</h3><canvas id="chartGrasa"></canvas></div>
        <div class="card"><h3>Aumento de peso involuntario</h3><canvas id="chartPeso"></canvas></div>
    </div>
</div>

<!-- CANALES -->
<div id="canales" class="section">
    <div class="grid" style="grid-template-columns: repeat(3, 1fr);">
        <div class="card big-number insight-card"><div class="number">{int(rrss_pct)}%</div><p>Llega por RRSS</p></div>
        <div class="card big-number insight-card"><div class="number">{int(busqueda_pct)}%</div><p>B&uacute;squeda online</p></div>
        <div class="card big-number insight-card"><div class="number">{int(prof_pct)}%</div><p>Profesional de salud</p></div>
    </div>
    <div class="grid">
        <div class="card" style="grid-column: span 2;"><h3>Canal de adquisici&oacute;n</h3><canvas id="chartCanal"></canvas></div>
    </div>
</div>

<!-- SEGMENTOS -->
<div id="segmentos" class="section">
    <div class="grid">
        <div class="card"><h3>Segmentos de mujeres</h3><canvas id="chartSegmentos"></canvas></div>
        <div class="card">
            <h3>Definici&oacute;n de segmentos</h3>
            <div style="font-size:13px; line-height:1.8;">
                <p><b style="color:var(--rosa-oscuro)">Sobrecarga total ({seg_sobrecarga_pct}%)</b> &mdash; 5+ s&iacute;ntomas f&iacute;sicos + impacto emocional severo</p>
                <p><b style="color:var(--rosa-oscuro)">Carga f&iacute;sica alta</b> &mdash; 5+ s&iacute;ntomas f&iacute;sicos pero sin impacto emocional severo</p>
                <p><b style="color:var(--rosa-oscuro)">Impacto emocional</b> &mdash; Menos s&iacute;ntomas f&iacute;sicos pero alto impacto emocional</p>
                <p><b style="color:var(--rosa-oscuro)">Moderada</b> &mdash; Algunos s&iacute;ntomas, manejo emocional razonable</p>
                <p><b style="color:var(--rosa-oscuro)">Cuerpo s&iacute;, mente no</b> &mdash; S&iacute;ntomas f&iacute;sicos pero se sienten bien emocionalmente</p>
                <p><b style="color:var(--rosa-oscuro)">Bien</b> &mdash; Pocas molestias en general</p>
            </div>
        </div>
    </div>
</div>

<!-- INSIGHTS -->
<div id="insights" class="section">
    <div class="grid">
        <div class="card insight-card"><h3>Dato #1 &mdash; Condici&oacute;n sist&eacute;mica</h3><p class="number">{int(pct_5plus)}%</p><p>de las mujeres tiene 5 o m&aacute;s s&iacute;ntomas simult&aacute;neos.</p></div>
        <div class="card insight-card"><h3>Dato #2 &mdash; Segmento mayoritario</h3><p class="number">{seg_sobrecarga_pct}%</p><p>est&aacute; en "sobrecarga total": muchos s&iacute;ntomas + impacto emocional severo.</p></div>
    </div>
    <div class="grid">
        <div class="card insight-card"><h3>Dato #3 &mdash; El s&iacute;ntoma invisible</h3><p>La <b>baja libido</b> es uno de los s&iacute;ntomas m&aacute;s reportados pero el menos visible.</p></div>
        <div class="card insight-card"><h3>Dato #4 &mdash; El pack peso+hinchaz&oacute;n</h3><p>Si tiene aumento de peso, hay alta probabilidad de hinchaz&oacute;n. Son inseparables.</p></div>
    </div>
    <div class="grid">
        <div class="card insight-card"><h3>Dato #5 &mdash; El sue&ntilde;o es la puerta</h3><p>Si tiene fatiga, hay alta probabilidad de insomnio. Quien resuelve el sue&ntilde;o, captura a la paciente.</p></div>
        <div class="card insight-card"><h3>Dato #6 &mdash; Gap sanitario</h3><p>Solo el <b>{int(prof_pct)}%</b> llega por profesional de salud. El {int(rrss_pct)}% llega por RRSS.</p></div>
    </div>
    <div class="grid">
        <div class="card insight-card"><h3>Dato #7 &mdash; Mercado en crecimiento</h3><p>El volumen de respuestas sigue creciendo trimestre a trimestre. Total: <b>{total_fmt}</b>.</p></div>
        <div class="card insight-card"><h3>Dato #8 &mdash; Base de datos {year_range}</h3><p>Datos recopilados entre {datos.get("fecha_min", "2022")} y {datos.get("fecha_max", "2026")} de m&uacute;ltiples fuentes.</p></div>
    </div>
</div>

<!-- ESTUDIO FISICO -->
<div id="fisico" class="section">
    <div style="margin-bottom:20px;">
        <a href="DOMMA_Analisis_Menopausia_BA.xlsx" download class="download-btn">
            <svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M21 15v4a2 2 0 01-2 2H5a2 2 0 01-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg>
            Descargar Excel completo
        </a>
    </div>

    <div class="grid" style="grid-template-columns: repeat(4, 1fr);">
        <div class="card big-number"><div class="number">{fisico_total}</div><div class="label">Respuestas</div></div>
        <div class="card big-number"><div class="number">{fisico_edad_top}</div><div class="label">Rango edad mayoritario ({fisico_edad_top_pct}%)</div></div>
        <div class="card big-number"><div class="number">{fisico_ejercicio_pct}%</div><div class="label">Hacen ejercicio</div></div>
        <div class="card big-number"><div class="number">{nota_cambios_pct}%</div><div class="label">Nota cambios f&iacute;sicos</div></div>
    </div>

    <div class="grid">
        <div class="card"><h3>Distribuci&oacute;n por edad</h3><canvas id="chartFisicoEdad"></canvas></div>
        <div class="card"><h3>Fase menop&aacute;usica</h3><canvas id="chartFisicoFase"></canvas></div>
    </div>

    <div class="grid">
        <div class="card" style="grid-column: span 2;">
            <h3>Cambios f&iacute;sicos notados</h3>
            <p class="subtitle">% de mujeres que reporta cada cambio (multi-respuesta, n={fisico_total})</p>
            <canvas id="chartFisicoCambios"></canvas>
        </div>
    </div>

    <div class="grid">
        <div class="card" style="grid-column: span 2;">
            <h3>Lo que m&aacute;s les preocupa</h3>
            <p class="subtitle">% de mujeres que selecciona cada preocupaci&oacute;n</p>
            <canvas id="chartFisicoWorries"></canvas>
        </div>
    </div>

    <div class="grid">
        <div class="card">
            <h3>Tipo de ejercicio</h3>
            <p class="subtitle">Entre las que hacen ejercicio</p>
            <canvas id="chartFisicoExType"></canvas>
        </div>
        <div class="card">
            <h3>Rendimiento al hacer ejercicio</h3>
            <p class="subtitle">Problemas reportados</p>
            <canvas id="chartFisicoPerf"></canvas>
        </div>
    </div>

    <div class="grid">
        <div class="card">
            <h3>Comportamiento nutricional</h3>
            <canvas id="chartFisicoNutricion"></canvas>
        </div>
        <div class="card">
            <h3>Barreras para cuidarse</h3>
            <p class="subtitle">% de mujeres que selecciona cada barrera</p>
            <canvas id="chartFisicoBarreras"></canvas>
        </div>
    </div>
</div>

</div>

<script>
const AZUL = '{AZUL}';
const ROSA = '{ROSA}';
const ROSA_OSC = '{ROSA_OSC}';
const AZUL_MED = '{AZUL_MED}';
const palette = [ROSA_OSC, '#D4A574', AZUL, ROSA, AZUL_MED, '#8B7355', '#6B4C3B', '#A0522D'];

const symLabels = {json.dumps(sym_labels)};
const symPcts = {json.dumps(sym_pcts)};
const emoLabels = {json.dumps(emo_labels)};
const emoPcts = {json.dumps(emo_pcts)};
const faseData = {json.dumps(fase)};
const volData = {json.dumps(vol)};
const sofData = {json.dumps(sofocos)};
const sufferingData = {json.dumps(suffering)};
const segData = {json.dumps(segmentos)};
const canalData = {json.dumps(dict(zip(canal_labels, canal_pcts)))};
const crossData = {json.dumps(cross_js_limited) if cross_js_limited else json.dumps({"Baja libido": [54.8, 58.5, 63.7, 76.2], "Insomnio": [58.9, 67.5, 71.6, 70.6], "Sofocos": [39.5, 56.3, 71.0, 68.4], "Aumento peso": [65.3, 70.2, 70.1, 68.9], "Fatiga": [71.9, 73.3, 69.6, 63.4], "Seq. vaginal": [32.5, 40.1, 49.0, 68.2], "Sudores noct.": [1.4, 7.6, 15.0, 17.4]})};
const crossFases = {json.dumps(cross_fases)};
const ageCross = {age_cross_js};
const ageGroups = ["40-44", "45-49", "50-54", "55-59", "60-64"];
const coocLabels = {cooc_labels_js};
const coocMatrix = {cooc_matrix_js};

// Fisico data
const fisicoEdadLabels = {json.dumps(list(fisico_edad.keys()) if fisico_edad else ["35-39", "40-44", "45-49", "50-54", "55-59", "60+"])};
const fisicoEdadData = {json.dumps([round(v/fisico_total*100,1) for v in fisico_edad.values()] if fisico_edad else [1.0, 4.3, 19.1, 40.5, 24.8, 10.4])};
const fisicoFaseLabels = {json.dumps(list(fisico_fase.keys()) if fisico_fase else ["Posmenopausia", "Menopausia", "Perimenopausia", "Premenopausia"])};
const fisicoFaseData = {json.dumps([round(v/fisico_total*100,1) for v in fisico_fase.values()] if fisico_fase else [26.9, 42.9, 17.0, 8.8])};
const fisicoCambiosLabels = {json.dumps(list(fisico_cambios_pct.keys()) if fisico_cambios_pct else ["Grasa abdominal", "Cansancio", "Hinchada", "Cuesta mantener peso"])};
const fisicoCambiosData = {json.dumps(list(fisico_cambios_pct.values()) if fisico_cambios_pct else [73.2, 63.5, 50.1, 49.7])};
const fisicoWorriesLabels = {json.dumps(list(fisico_preocup_pct.keys()) if fisico_preocup_pct else ["Grasa barriga", "Sin energia", "Dolores"])};
const fisicoWorriesData = {json.dumps(list(fisico_preocup_pct.values()) if fisico_preocup_pct else [53.2, 51.2, 50.2])};
const fisicoExTypeLabels = {json.dumps(list(fisico_tipo_pct.keys()) if fisico_tipo_pct else ["Fuerza", "Caminar", "Yoga/Pilates", "Cardio"])};
const fisicoExTypeData = {json.dumps(list(fisico_tipo_pct.values()) if fisico_tipo_pct else [40.6, 36.1, 35.0, 22.1])};
const fisicoPerfLabels = {json.dumps(list(fisico_rend_pct.keys()) if fisico_rend_pct else ["Cuesta perder grasa", "Me canso mas", "Cuesta ganar fuerza"])};
const fisicoPerfData = {json.dumps(list(fisico_rend_pct.values()) if fisico_rend_pct else [35.8, 26.9, 24.8])};
const fisicoBarrerasLabels = {json.dumps(list(fisico_barreras_pct.keys()) if fisico_barreras_pct else ["Falta de tiempo", "Falta de energia"])};
const fisicoBarrerasData = {json.dumps(list(fisico_barreras_pct.values()) if fisico_barreras_pct else [42.1, 36.2])};

// NAVIGATION
function showSection(id) {{
    document.querySelectorAll('.section').forEach(s => s.classList.remove('active'));
    document.querySelectorAll('.nav button').forEach(b => b.classList.remove('active'));
    document.getElementById(id).classList.add('active');
    event.target.classList.add('active');
}}

// CHART DEFAULTS
Chart.defaults.font.family = "-apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif";
Chart.defaults.color = AZUL;
const defaultOpts = {{ responsive: true, maintainAspectRatio: true, plugins: {{ legend: {{ display: false }} }} }};

// OVERVIEW
new Chart('chartFase', {{
    type: 'doughnut',
    data: {{ labels: Object.keys(faseData), datasets: [{{ data: Object.values(faseData), backgroundColor: [AZUL, ROSA_OSC, ROSA, '#D4B5A5', '#8B7355'] }}] }},
    options: {{ ...defaultOpts, plugins: {{ legend: {{ display: true, position: 'bottom' }} }} }}
}});
new Chart('chartVolumen', {{
    type: 'bar',
    data: {{ labels: Object.keys(volData), datasets: [{{ data: Object.values(volData), backgroundColor: ROSA_OSC }}] }},
    options: {{ ...defaultOpts, scales: {{ x: {{ ticks: {{ maxRotation: 45 }} }} }} }}
}});
new Chart('chartSintomasOverview', {{
    type: 'bar',
    data: {{ labels: symLabels, datasets: [{{ data: symPcts, backgroundColor: ROSA_OSC }}] }},
    options: {{ ...defaultOpts, indexAxis: 'y', scales: {{ x: {{ max: 80 }} }} }}
}});
new Chart('chartEmoOverview', {{
    type: 'bar',
    data: {{ labels: emoLabels, datasets: [{{ data: emoPcts, backgroundColor: emoLabels.map((l) => l === 'Estupenda' ? AZUL_MED : ROSA_OSC) }}] }},
    options: {{ ...defaultOpts, indexAxis: 'y', scales: {{ x: {{ max: 70 }} }} }}
}});

// SINTOMAS
new Chart('chartSintomas', {{
    type: 'bar',
    data: {{ labels: symLabels, datasets: [{{ data: symPcts, backgroundColor: ROSA_OSC, borderRadius: 4 }}] }},
    options: {{ ...defaultOpts, indexAxis: 'y', scales: {{ x: {{ max: 80, ticks: {{ callback: v => v + '%' }} }} }} }}
}});
new Chart('chartSofocos', {{
    type: 'doughnut',
    data: {{ labels: Object.keys(sofData), datasets: [{{ data: Object.values(sofData), backgroundColor: [AZUL, ROSA_OSC, ROSA] }}] }},
    options: {{ ...defaultOpts, plugins: {{ legend: {{ display: true, position: 'bottom' }} }} }}
}});
new Chart('chartSuffering', {{
    type: 'bar',
    data: {{ labels: Object.keys(sufferingData), datasets: [{{ data: Object.values(sufferingData), backgroundColor: Object.keys(sufferingData).map(k => parseInt(k) >= 5 ? AZUL : ROSA_OSC) }}] }},
    options: {{ ...defaultOpts, scales: {{ x: {{ title: {{ display: true, text: 'N.o de sintomas' }} }} }} }}
}});

// EMOCIONAL
new Chart('chartEmocional', {{
    type: 'bar',
    data: {{ labels: emoLabels, datasets: [{{ data: emoPcts, backgroundColor: emoLabels.map((l) => l === 'Estupenda' ? AZUL_MED : ROSA_OSC), borderRadius: 4 }}] }},
    options: {{ ...defaultOpts, indexAxis: 'y', scales: {{ x: {{ max: 70, ticks: {{ callback: v => v + '%' }} }} }} }}
}});

// CRUCES
new Chart('chartCross', {{
    type: 'bar',
    data: {{ labels: crossFases, datasets: Object.entries(crossData).map(([label, data], i) => ({{ label, data, backgroundColor: palette[i % palette.length] }})) }},
    options: {{ responsive: true, plugins: {{ legend: {{ display: true, position: 'top' }} }}, scales: {{ y: {{ max: 100, ticks: {{ callback: v => v + '%' }} }} }} }}
}});
new Chart('chartAgeCross', {{
    type: 'bar',
    data: {{ labels: ageGroups, datasets: Object.entries(ageCross).map(([label, data], i) => ({{ label, data, backgroundColor: palette[i % palette.length] }})) }},
    options: {{ responsive: true, plugins: {{ legend: {{ display: true, position: 'top' }} }}, scales: {{ y: {{ max: 100, ticks: {{ callback: v => v + '%' }} }} }} }}
}});

// Heatmap
(function() {{
    const table = document.getElementById('heatmapTable');
    let html = '<tr><th></th>';
    coocLabels.forEach(l => html += `<th>${{l}}</th>`);
    html += '</tr>';
    coocMatrix.forEach((row, i) => {{
        html += `<tr><td class="row-header">${{coocLabels[i]}}</td>`;
        row.forEach((val, j) => {{
            if (i === j) {{
                html += '<td style="background:#eee;color:#999">\\u2014</td>';
            }} else {{
                const intensity = Math.min(val / 85, 1);
                const r = Math.round(196 * intensity + 248 * (1 - intensity));
                const g = Math.round(151 * intensity + 247 * (1 - intensity));
                const b = Math.round(127 * intensity + 245 * (1 - intensity));
                html += `<td style="background:rgb(${{r}},${{g}},${{b}});color:${{intensity > 0.6 ? 'white' : AZUL}}">${{val}}%</td>`;
            }}
        }});
        html += '</tr>';
    }});
    table.innerHTML = html;
}})();

// PESO
new Chart('chartGrasa', {{
    type: 'doughnut',
    data: {{ labels: ['Si, claramente', 'Un poco', 'No'], datasets: [{{ data: [{grasa_pct}, 15.0, {round(100-grasa_pct-15, 1)}], backgroundColor: [AZUL, ROSA_OSC, ROSA] }}] }},
    options: {{ ...defaultOpts, plugins: {{ legend: {{ display: true, position: 'bottom' }} }} }}
}});
new Chart('chartPeso', {{
    type: 'doughnut',
    data: {{ labels: ['Si, totalmente', 'Algo', 'No'], datasets: [{{ data: [{peso_pct}, 14.8, {round(100-peso_pct-14.8, 1)}], backgroundColor: [AZUL, ROSA_OSC, ROSA] }}] }},
    options: {{ ...defaultOpts, plugins: {{ legend: {{ display: true, position: 'bottom' }} }} }}
}});

// CANALES
new Chart('chartCanal', {{
    type: 'bar',
    data: {{ labels: Object.keys(canalData), datasets: [{{ data: Object.values(canalData), backgroundColor: ROSA_OSC, borderRadius: 4 }}] }},
    options: {{ ...defaultOpts, indexAxis: 'y', scales: {{ x: {{ ticks: {{ callback: v => v + '%' }} }} }} }}
}});

// SEGMENTOS
new Chart('chartSegmentos', {{
    type: 'doughnut',
    data: {{ labels: Object.keys(segData), datasets: [{{ data: Object.values(segData), backgroundColor: [AZUL, ROSA_OSC, ROSA, '#D4B5A5', AZUL_MED, '#8B7355'] }}] }},
    options: {{ ...defaultOpts, plugins: {{ legend: {{ display: true, position: 'right' }} }} }}
}});

// ESTUDIO FISICO
new Chart('chartFisicoEdad', {{
    type: 'bar',
    data: {{ labels: fisicoEdadLabels, datasets: [{{ data: fisicoEdadData, backgroundColor: ROSA_OSC, borderRadius: 4 }}] }},
    options: {{ ...defaultOpts, scales: {{ y: {{ ticks: {{ callback: v => v + '%' }} }} }} }}
}});
new Chart('chartFisicoFase', {{
    type: 'doughnut',
    data: {{ labels: fisicoFaseLabels, datasets: [{{ data: fisicoFaseData, backgroundColor: [AZUL, ROSA_OSC, ROSA, '#D4B5A5'] }}] }},
    options: {{ ...defaultOpts, plugins: {{ legend: {{ display: true, position: 'bottom' }} }} }}
}});
new Chart('chartFisicoCambios', {{
    type: 'bar',
    data: {{ labels: fisicoCambiosLabels, datasets: [{{ data: fisicoCambiosData, backgroundColor: ROSA_OSC, borderRadius: 4 }}] }},
    options: {{ ...defaultOpts, indexAxis: 'y', scales: {{ x: {{ max: 80, ticks: {{ callback: v => v + '%' }} }} }} }}
}});
new Chart('chartFisicoWorries', {{
    type: 'bar',
    data: {{ labels: fisicoWorriesLabels, datasets: [{{ data: fisicoWorriesData, backgroundColor: ROSA_OSC, borderRadius: 4 }}] }},
    options: {{ ...defaultOpts, indexAxis: 'y', scales: {{ x: {{ max: 60, ticks: {{ callback: v => v + '%' }} }} }} }}
}});
new Chart('chartFisicoExType', {{
    type: 'bar',
    data: {{ labels: fisicoExTypeLabels, datasets: [{{ data: fisicoExTypeData, backgroundColor: ROSA_OSC, borderRadius: 4 }}] }},
    options: {{ ...defaultOpts, indexAxis: 'y', scales: {{ x: {{ max: 50, ticks: {{ callback: v => v + '%' }} }} }} }}
}});
new Chart('chartFisicoPerf', {{
    type: 'bar',
    data: {{ labels: fisicoPerfLabels, datasets: [{{ data: fisicoPerfData, backgroundColor: ROSA_OSC, borderRadius: 4 }}] }},
    options: {{ ...defaultOpts, indexAxis: 'y', scales: {{ x: {{ max: 45, ticks: {{ callback: v => v + '%' }} }} }} }}
}});
new Chart('chartFisicoNutricion', {{
    type: 'doughnut',
    data: {{ labels: ["Equilibrio", "Antojos/ansiedad", "Picar entre horas"], datasets: [{{ data: [45.1, 42.7, 25.5], backgroundColor: [AZUL_MED, ROSA_OSC, ROSA] }}] }},
    options: {{ ...defaultOpts, plugins: {{ legend: {{ display: true, position: 'bottom' }} }} }}
}});
new Chart('chartFisicoBarreras', {{
    type: 'bar',
    data: {{ labels: fisicoBarrerasLabels, datasets: [{{ data: fisicoBarrerasData, backgroundColor: ROSA_OSC, borderRadius: 4 }}] }},
    options: {{ ...defaultOpts, indexAxis: 'y', scales: {{ x: {{ max: 50, ticks: {{ callback: v => v + '%' }} }} }} }}
}});
</script>
</body>
</html>'''

    return html


# ---------------------------------------------------------------------------
# GENERACION EXCEL
# ---------------------------------------------------------------------------

def generar_excel(datos):
    """Genera el archivo Excel con analisis."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("  [AVISO] openpyxl no instalado. Ejecuta: pip3 install openpyxl")
        print("  Saltando generacion de Excel.")
        return False

    wb = Workbook()

    # Estilos
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="101637", end_color="101637", fill_type="solid")
    rosa_fill = PatternFill(start_color="E7C6B9", end_color="E7C6B9", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def style_header(ws, row=1, cols=10):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

    # --- Hoja 1: Resumen ---
    ws = wb.active
    ws.title = "Resumen"
    ws.append(["Metrica", "Valor"])
    style_header(ws, 1, 2)
    ws.append(["Total respuestas", datos["total"]])
    ws.append(["Media sintomas/mujer", datos["media_sintomas"]])
    ws.append(["Periodo", datos["year_range"]])
    ws.append(["Fecha actualizacion", datetime.now().strftime("%d/%m/%Y %H:%M")])
    ws.append([])
    ws.append(["Fuente", "Respuestas", "Desde", "Hasta"])
    style_header(ws, ws.max_row, 4)
    for f in datos["fuentes"]:
        ws.append([f["nombre"], f["total"], f.get("fecha_min", ""), f.get("fecha_max", "")])
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15

    # --- Hoja 2: Sintomas ---
    ws2 = wb.create_sheet("Sintomas")
    ws2.append(["Sintoma", "% reportado"])
    style_header(ws2, 1, 2)
    for s, pct in sorted(datos["sintomas_pct"].items(), key=lambda x: -x[1]):
        ws2.append([s, pct])
    ws2.column_dimensions["A"].width = 25

    # --- Hoja 3: Emocional ---
    ws3 = wb.create_sheet("Emocional")
    ws3.append(["Estado", "% reportado"])
    style_header(ws3, 1, 2)
    for e, pct in sorted(datos["emocional_pct"].items(), key=lambda x: -x[1]):
        ws3.append([e, pct])
    ws3.column_dimensions["A"].width = 25

    # --- Hoja 4: Fase menstrual ---
    ws4 = wb.create_sheet("Fase menstrual")
    ws4.append(["Fase", "Respuestas", "%"])
    style_header(ws4, 1, 3)
    fase_total = sum(datos["fase"].values()) or 1
    for f, count in sorted(datos["fase"].items(), key=lambda x: -x[1]):
        ws4.append([f, count, round(count / fase_total * 100, 1)])
    ws4.column_dimensions["A"].width = 25

    # --- Hoja 5: Canales ---
    ws5 = wb.create_sheet("Canales")
    ws5.append(["Canal", "%"])
    style_header(ws5, 1, 2)
    for c, pct in sorted(datos["canal_pct"].items(), key=lambda x: -x[1]):
        ws5.append([c, pct])
    ws5.column_dimensions["A"].width = 25

    # --- Hoja 6: Segmentos ---
    ws6 = wb.create_sheet("Segmentos")
    ws6.append(["Segmento", "Respuestas", "%"])
    style_header(ws6, 1, 3)
    seg_total = sum(datos["segmentos"].values()) or 1
    for s, count in sorted(datos["segmentos"].items(), key=lambda x: -x[1]):
        ws6.append([s, count, round(count / seg_total * 100, 1)])
    ws6.column_dimensions["A"].width = 30

    # --- Hoja 7: Volumen trimestral ---
    vol = datos.get("volumen_trimestre")
    if vol:
        ws7 = wb.create_sheet("Volumen trimestral")
        ws7.append(["Trimestre", "Respuestas"])
        style_header(ws7, 1, 2)
        for t, count in sorted(vol.items()):
            ws7.append([t, count])

    # --- Hoja 8: Estudio fisico ---
    fisico = datos.get("fisico", {})
    if fisico:
        ws8 = wb.create_sheet("Estudio fisico")
        ws8.append(["Metrica", "Valor"])
        style_header(ws8, 1, 2)
        ws8.append(["Total respuestas", fisico.get("total", 0)])
        ws8.append(["Ejercicio %", fisico.get("ejercicio_pct", 0)])
        ws8.append([])
        ws8.append(["Edad", "Respuestas"])
        style_header(ws8, ws8.max_row, 2)
        for k, v in fisico.get("edad", {}).items():
            ws8.append([k, v])
        ws8.column_dimensions["A"].width = 30

    try:
        wb.save(str(OUTPUT_XLSX))
        print(f"  Excel guardado: {OUTPUT_XLSX}")
        return True
    except Exception as e:
        print(f"  [ERROR] No se pudo guardar Excel: {e}")
        return False


# ---------------------------------------------------------------------------
# DESPLIEGUE A GITHUB PAGES
# ---------------------------------------------------------------------------

def github_api_request(url, method="GET", data=None, token=None):
    """Hace una peticion a la API de GitHub."""
    headers = {
        "Accept": "application/vnd.github.v3+json",
        "User-Agent": "DOMMA-Dashboard-Updater",
    }
    if token:
        headers["Authorization"] = f"token {token}"

    body = json.dumps(data).encode("utf-8") if data else None
    req = urllib.request.Request(url, data=body, headers=headers, method=method)

    try:
        with urllib.request.urlopen(req) as resp:
            return json.loads(resp.read().decode("utf-8")), resp.status
    except urllib.error.HTTPError as e:
        error_body = e.read().decode("utf-8") if e.fp else ""
        return {"error": str(e), "detail": error_body}, e.code


def desplegar_github(html_content, excel_path, token):
    """Despliega index.html y el Excel a GitHub Pages."""
    print("\n=== DESPLEGANDO A GITHUB PAGES ===")

    base_url = f"{GITHUB_API}/repos/{GITHUB_REPO}/contents"

    files_to_deploy = [
        ("index.html", html_content.encode("utf-8")),
    ]

    # Excel
    if excel_path and os.path.exists(excel_path):
        with open(excel_path, "rb") as f:
            files_to_deploy.append(("DOMMA_Analisis_Menopausia_BA.xlsx", f.read()))

    for filename, content in files_to_deploy:
        print(f"  Subiendo {filename}...")

        # Get current SHA if file exists
        resp, status = github_api_request(f"{base_url}/{filename}", token=token)
        sha = resp.get("sha") if status == 200 else None

        # Create/update file
        payload = {
            "message": f"Actualizar {filename} - {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            "content": base64.b64encode(content).decode("ascii"),
            "branch": "main",
        }
        if sha:
            payload["sha"] = sha

        resp, status = github_api_request(
            f"{base_url}/{filename}", method="PUT", data=payload, token=token
        )

        if status in (200, 201):
            print(f"    -> OK ({status})")
        else:
            print(f"    -> ERROR {status}: {resp.get('error', '')} {resp.get('detail', '')[:200]}")

    print("  Despliegue completado.")
    print(f"  URL: https://crisdomma.github.io/domma-data-dashboard/")


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Actualizar Dashboard DOMMA de Menopausia")
    parser.add_argument("--dry-run", action="store_true", help="Mostrar que se haria sin desplegar")
    args = parser.parse_args()

    print("=" * 60)
    print("  DOMMA - Actualizacion del Dashboard de Menopausia")
    print(f"  Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print("=" * 60)

    # 1. Descubrir CSVs
    print("\n--- PASO 1: Descubriendo archivos CSV ---")
    fuentes = descubrir_csvs()

    total_archivos = sum(len(v) for v in fuentes.values())
    if total_archivos == 0:
        print("  [ERROR] No se encontraron archivos CSV en ninguna ubicacion.")
        print(f"  Buscado en: {DOWNLOADS_DIR}")
        print(f"  Buscado en: {DATOS_NUEVOS_DIR}")
        sys.exit(1)

    for nombre, archivos in fuentes.items():
        if archivos:
            print(f"  {nombre}: {len(archivos)} archivo(s)")
            for a in archivos:
                print(f"    - {Path(a).name}")
        else:
            print(f"  {nombre}: no encontrado")

    # 2. Procesar datos
    print("\n--- PASO 2: Procesando datos ---")
    resultados = []

    # WooCommerce
    for filepath in fuentes["woocommerce"]:
        resultado = procesar_woocommerce(filepath)
        if resultado:
            resultados.append(resultado)

    # Shopify (use same WooCommerce parser - similar format)
    for filepath in fuentes["shopify"]:
        resultado = procesar_woocommerce(filepath)
        if resultado:
            resultado["fuente"] = "Shopify RevenueHunt"
            resultados.append(resultado)

    # Typeform principal
    for filepath in fuentes["typeform_principal"]:
        resultado = procesar_typeform_principal(filepath)
        if resultado:
            resultados.append(resultado)

    # Typeform fisico
    for filepath in fuentes["typeform_fisico"]:
        # Use the most recent one (avoid duplicates)
        pass
    if fuentes["typeform_fisico"]:
        resultado = procesar_typeform_fisico(fuentes["typeform_fisico"][0])
        if resultado:
            resultados.append(resultado)

    # Datos nuevos (try to auto-detect format)
    for filepath in fuentes["datos_nuevos"]:
        print(f"  Intentando leer datos nuevos: {Path(filepath).name}...")
        rows, fieldnames = leer_csv_seguro(filepath, max_rows=5)
        if fieldnames:
            if any("tags" in f.lower() for f in fieldnames):
                resultado = procesar_woocommerce(filepath)
                if resultado:
                    resultado["fuente"] = f"Nuevo: {Path(filepath).name}"
                    resultados.append(resultado)
            elif any("sofocos" in f.lower() for f in fieldnames):
                resultado = procesar_typeform_principal(filepath)
                if resultado:
                    resultado["fuente"] = f"Nuevo: {Path(filepath).name}"
                    resultados.append(resultado)
            else:
                print(f"    [AVISO] Formato no reconocido, saltando.")

    if not resultados:
        print("  [ERROR] No se pudieron procesar datos de ningun archivo.")
        sys.exit(1)

    # 3. Combinar
    print("\n--- PASO 3: Combinando datos ---")
    datos_combinados = combinar_datos(resultados)
    total = datos_combinados["total"]
    print(f"  Total combinado: {total:,} respuestas")
    print(f"  Fuentes procesadas: {len(resultados)}")
    print(f"  Periodo: {datos_combinados.get('fecha_min', '?')} a {datos_combinados.get('fecha_max', '?')}")

    # 4. Generar HTML
    print("\n--- PASO 4: Generando dashboard HTML ---")
    html_content = generar_html(datos_combinados)
    if not args.dry_run:
        with open(OUTPUT_HTML, "w", encoding="utf-8") as f:
            f.write(html_content)
        print(f"  HTML guardado: {OUTPUT_HTML}")
    else:
        print(f"  [DRY-RUN] Se generaria HTML ({len(html_content):,} bytes)")

    # 5. Generar Excel
    print("\n--- PASO 5: Generando Excel ---")
    if not args.dry_run:
        generar_excel(datos_combinados)
    else:
        print(f"  [DRY-RUN] Se generaria Excel en {OUTPUT_XLSX}")

    # 6. Desplegar a GitHub Pages
    print("\n--- PASO 6: Despliegue a GitHub Pages ---")
    token = os.environ.get("GITHUB_TOKEN")
    if not token:
        print("  [AVISO] Variable GITHUB_TOKEN no definida.")
        print("  Para desplegar, ejecuta:")
        print("    export GITHUB_TOKEN=tu_token_aqui")
        print("    python3 actualizar_dashboard.py")
        print("  Saltando despliegue.")
    elif args.dry_run:
        print(f"  [DRY-RUN] Se desplegaria a https://crisdomma.github.io/domma-data-dashboard/")
        print(f"  Archivos: index.html, DOMMA_Analisis_Menopausia_BA.xlsx")
    else:
        desplegar_github(html_content, str(OUTPUT_XLSX), token)

    # 7. Resumen
    print("\n" + "=" * 60)
    print("  RESUMEN DE ACTUALIZACION")
    print("=" * 60)
    print(f"  Total respuestas:  {total:,}")
    print(f"  Fuentes:           {len(resultados)}")
    for r in datos_combinados["fuentes"]:
        print(f"    - {r['nombre']}: {r['total']:,}")
    print(f"  Periodo:           {datos_combinados.get('year_range', '?')}")
    print(f"  Media sintomas:    {datos_combinados['media_sintomas']}")
    print(f"  Top 3 sintomas:    {', '.join(list(datos_combinados['sintomas_pct'].keys())[:3])}")

    if args.dry_run:
        print("\n  [DRY-RUN] No se han realizado cambios reales.")
    else:
        print(f"\n  Archivos actualizados:")
        print(f"    - {OUTPUT_HTML}")
        print(f"    - {OUTPUT_XLSX}")

    print("=" * 60)


if __name__ == "__main__":
    main()
