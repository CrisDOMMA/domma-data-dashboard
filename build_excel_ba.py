#!/usr/bin/env python3
"""
DOMMA Power BA Excel Workbook Generator
Reads 4 CSV sources and creates a professional multi-sheet Excel workbook.
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo
import warnings
import re
from datetime import datetime

warnings.filterwarnings('ignore')

# --- PATHS ---
WOOCOMMERCE_PATH = "/Users/cristina/Downloads/jyHJgz-2026-03-24-640823c5-e7b5-4989-82b0-d6bf5189b17d.csv"
TYPEFORM_PATH = "/Users/cristina/Downloads/responses-rXKhc6VM-01KMMNNV4GKZJJ74CQTGMST13N-LLKEXQMFKH8KU0DCFUB7MLR2.csv"
TYPEFORM_PHYS_PATH = "/Users/cristina/Downloads/responses-fGe6BPkD-01KMMWP887WXBP4VQ5MPST1YCB-4NS7LO8GFSG7B2BAKIQR5JQ1.csv"
SHOPIFY_PATH = "/Users/cristina/Documents/DEVELOPER/DOMMA/domma-data-dashboard/datos_nuevos/DOMMA_Quiz_responses_2026-03-30_142714.csv"
OUTPUT_PATH = "/Users/cristina/Documents/DEVELOPER/DOMMA/domma-data-dashboard/DOMMA_Analisis_Menopausia_BA.xlsx"

# --- COLORS ---
DARK_BLUE = "101637"
PINK = "E7C6B9"
CREAM = "F8F7F5"
WHITE = "FFFFFF"

# --- STYLES ---
header_font = Font(name='Calibri', bold=True, color=WHITE, size=11)
header_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type='solid')
pink_fill = PatternFill(start_color=PINK, end_color=PINK, fill_type='solid')
cream_fill = PatternFill(start_color=CREAM, end_color=CREAM, fill_type='solid')
white_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type='solid')
title_font = Font(name='Calibri', bold=True, color=DARK_BLUE, size=16)
subtitle_font = Font(name='Calibri', bold=True, color=DARK_BLUE, size=13)
metric_font = Font(name='Calibri', bold=True, color=DARK_BLUE, size=24)
metric_label_font = Font(name='Calibri', color="666666", size=10)
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

# --- STANDARD SYMPTOMS ---
STANDARD_SYMPTOMS = [
    'Baja libido', 'Insomnio', 'Sofocos', 'Aumento peso', 'Sequedad',
    'Fatiga', 'Dolor articular', 'Hinchazón', 'Caída pelo',
    'Sudores nocturnos', 'Problemas sueño', 'Cansancio'
]

EMOTIONAL_COLS = ['Ansiosa', 'Estresada', 'Cambios_humor', 'Irritable', 'Triste', 'Descentrada']

# --- SYMPTOM MAPPING ---
SYMPTOM_MAP = {
    'baja libido': 'Baja libido',
    'libido': 'Baja libido',
    'insomnio': 'Insomnio',
    'sofocos': 'Sofocos',
    'aumento peso': 'Aumento peso',
    'aumento de peso': 'Aumento peso',
    'subida de peso': 'Aumento peso',
    'peso': 'Aumento peso',
    'sequedad': 'Sequedad',
    'sequedad vaginal': 'Sequedad',
    'sequedad externa': 'Sequedad',
    'fatiga': 'Fatiga',
    'dolor articular': 'Dolor articular',
    'dolor de articulaciones': 'Dolor articular',
    'hinchazón': 'Hinchazón',
    'hinchaz': 'Hinchazón',
    'caída pelo': 'Caída pelo',
    'caida pelo': 'Caída pelo',
    'caída de pelo': 'Caída pelo',
    'pelo': 'Caída pelo',
    'sudores nocturnos': 'Sudores nocturnos',
    'sudores': 'Sudores nocturnos',
    'problemas de sueño': 'Problemas sueño',
    'problemas sueño': 'Problemas sueño',
    'depertares nocturnos': 'Problemas sueño',
    'despertares nocturnos': 'Problemas sueño',
    'cansancio': 'Cansancio',
    'pérdida de memoria': 'Pérdida memoria',
    'perdida de memoria': 'Pérdida memoria',
    'niebla mental': 'Niebla mental',
    'cambios de humor': 'Cambios humor',
    'ansiedad': 'Ansiedad',
    'irritabilidad': 'Irritabilidad',
}

def map_symptom(raw):
    """Map a raw symptom string to standard name."""
    if pd.isna(raw):
        return None
    raw_lower = str(raw).strip().lower()
    # Direct match
    if raw_lower in SYMPTOM_MAP:
        return SYMPTOM_MAP[raw_lower]
    # Partial match
    for key, val in SYMPTOM_MAP.items():
        if key in raw_lower:
            return val
    return str(raw).strip()

def parse_symptoms_slash(text):
    """Parse slash or comma+y separated symptom strings."""
    if pd.isna(text) or str(text).strip() == '':
        return []
    text = str(text)
    # Remove extra quotes
    text = text.strip('"').strip("'").strip()
    # Split by / or , or " y "
    parts = re.split(r'\s*/\s*|\s*,\s*|\s+y\s+', text)
    symptoms = []
    for p in parts:
        p = p.strip().strip('"').strip()
        if p:
            mapped = map_symptom(p)
            if mapped:
                symptoms.append(mapped)
    return list(set(symptoms))

def strip_shopify_quotes(val):
    """Strip extra quotes from Shopify values."""
    if pd.isna(val):
        return val
    s = str(val).strip()
    while s.startswith('"') and s.endswith('"') and len(s) > 1:
        s = s[1:-1].strip()
    return s

def normalize_fase(raw):
    """Normalize menstrual phase to standard categories."""
    if pd.isna(raw):
        return 'Sin datos'
    s = str(raw).strip().upper()
    if 'REGULAR' in s:
        return 'Regular'
    elif 'NADA' in s or 'NI ESTÁ' in s:
        return 'Menopausia'
    elif 'ALTERADA' in s:
        return 'Alterada'
    elif 'POCO' in s or 'FRECUENTE' in s:
        return 'Poco frecuente'
    elif 'PERIMENO' in s:
        return 'Perimenopausia'
    elif 'POST' in s:
        return 'Postmenopausia'
    elif 'PRE' in s:
        return 'Premenopausia'
    else:
        return str(raw).strip()[:30]

def normalize_canal(raw):
    """Normalize acquisition channel."""
    if pd.isna(raw):
        return 'Sin datos'
    s = str(raw).strip().lower()
    if 'instagram' in s or 'redes' in s or 'domma.stories' in s or 'social' in s:
        return 'Redes sociales'
    elif 'google' in s or 'buscador' in s:
        return 'Google/Buscador'
    elif 'amiga' in s or 'conocida' in s or 'recomend' in s or 'boca' in s:
        return 'Recomendación'
    elif 'médic' in s or 'medic' in s or 'profesional' in s or 'ginec' in s:
        return 'Profesional salud'
    elif 'podcast' in s:
        return 'Podcast'
    elif 'prensa' in s or 'medio' in s or 'tv' in s or 'tele' in s:
        return 'Medios/Prensa'
    elif 'facebook' in s or 'meta' in s:
        return 'Facebook'
    elif 'tiktok' in s:
        return 'TikTok'
    elif 'farmacia' in s:
        return 'Farmacia'
    elif 'otro' in s:
        return 'Otro'
    else:
        return str(raw).strip()[:30]

def parse_emotional_slash(text):
    """Parse emotional states from slash-separated string."""
    if pd.isna(text) or str(text).strip() == '':
        return {}
    text = str(text).strip('"').strip("'").strip()
    parts = re.split(r'\s*/\s*|\s*,\s*|\s+y\s+', text)
    result = {e: 0 for e in EMOTIONAL_COLS}
    for p in parts:
        p_lower = p.strip().lower()
        if 'ansio' in p_lower:
            result['Ansiosa'] = 1
        if 'estres' in p_lower:
            result['Estresada'] = 1
        if 'cambio' in p_lower or 'humor' in p_lower:
            result['Cambios_humor'] = 1
        if 'irrit' in p_lower:
            result['Irritable'] = 1
        if 'trist' in p_lower:
            result['Triste'] = 1
        if 'descen' in p_lower:
            result['Descentrada'] = 1
    return result

# ========================================================
# 1. LOAD AND PROCESS DATA
# ========================================================
print("Loading WooCommerce data...")
woo_cols_to_use = [0, 3, 6, 17, 19, 20, 22, 23, 26]
woo_raw = pd.read_csv(WOOCOMMERCE_PATH, usecols=woo_cols_to_use, low_memory=False)
woo_all_cols = pd.read_csv(WOOCOMMERCE_PATH, nrows=0).columns.tolist()
print(f"  WooCommerce columns loaded: {len(woo_raw.columns)}, rows: {len(woo_raw)}")

# Map by position name
woo = pd.DataFrame()
woo['Fuente'] = 'WooCommerce'
woo['Fecha'] = pd.to_datetime(woo_raw.iloc[:, 0], errors='coerce', utc=True)
# Find columns by partial name match from loaded columns
col_names = woo_raw.columns.tolist()
print(f"  WooCommerce loaded col names: {col_names}")

# Map columns by their content
fecha_col = [c for c in col_names if 'created_at' in c.lower()]
symptoms_col = [c for c in col_names if 'síntomas' in c.lower() or 'sintomas' in c.lower()]
fase_col = [c for c in col_names if 'menstruación' in c.lower() or 'menstruacion' in c.lower()]
canal_col = [c for c in col_names if 'conocido' in c.lower()]
emocional_col = [c for c in col_names if 'mental' in c.lower() or 'emocional' in c.lower()]
sexual_col = [c for c in col_names if 'sexual' in c.lower() or 'íntima' in c.lower()]
sofocos_col = [c for c in col_names if 'sofocos' in c.lower() and 'sudores' in c.lower()]
country_col = [c for c in col_names if 'country' in c.lower()]
device_col = [c for c in col_names if 'device' in c.lower()]

woo['Fecha'] = pd.to_datetime(woo_raw[fecha_col[0]], errors='coerce', utc=True) if fecha_col else pd.NaT
woo['Symptoms_raw'] = woo_raw[symptoms_col[0]] if symptoms_col else ''
woo['Fase_raw'] = woo_raw[fase_col[0]] if fase_col else ''
woo['Canal_raw'] = woo_raw[canal_col[0]] if canal_col else ''
woo['Emocional_raw'] = woo_raw[emocional_col[0]] if emocional_col else ''
woo['Sexual_raw'] = woo_raw[sexual_col[0]] if sexual_col else ''
woo['Sofocos_cuando'] = woo_raw[sofocos_col[0]] if sofocos_col else ''
woo['Country'] = woo_raw[country_col[0]] if country_col else ''
woo['Device'] = woo_raw[device_col[0]] if device_col else ''

# Parse symptoms
print("  Parsing WooCommerce symptoms...")
woo['Symptoms_list'] = woo['Symptoms_raw'].apply(parse_symptoms_slash)
for s in STANDARD_SYMPTOMS:
    woo[s] = woo['Symptoms_list'].apply(lambda x: 1 if s in x else 0)
woo['Num_Sintomas'] = woo[STANDARD_SYMPTOMS].sum(axis=1)

# Parse emotional
woo_emo = woo['Emocional_raw'].apply(parse_emotional_slash)
emo_df = pd.DataFrame(woo_emo.tolist())
for e in EMOTIONAL_COLS:
    woo[e] = emo_df[e] if e in emo_df.columns else 0

woo['Fase_Menstrual'] = woo['Fase_raw'].apply(normalize_fase)
woo['Canal'] = woo['Canal_raw'].apply(normalize_canal)

del woo_raw
print(f"  WooCommerce processed: {len(woo)} rows")

# --- SHOPIFY ---
print("Loading Shopify data...")
shop_raw = pd.read_csv(SHOPIFY_PATH, low_memory=False)
print(f"  Shopify columns: {len(shop_raw.columns)}, rows: {len(shop_raw)}")

shop = pd.DataFrame()
shop['Fuente'] = 'Shopify'

# Find columns
shop_cols = shop_raw.columns.tolist()
fecha_s = [c for c in shop_cols if 'created at' in c.lower()]
symptoms_s = [c for c in shop_cols if 'q8' in c.lower() and 'síntoma' in c.lower()]
fase_s = [c for c in shop_cols if 'q3' in c.lower() and 'menstruación' in c.lower()]
canal_s = [c for c in shop_cols if 'q16' in c.lower() and 'conocido' in c.lower()]
emocional_s = [c for c in shop_cols if 'q11' in c.lower() and ('mental' in c.lower() or 'emocional' in c.lower())]
sexual_s = [c for c in shop_cols if 'q12' in c.lower() and ('sexual' in c.lower() or 'íntima' in c.lower())]
sofocos_s = [c for c in shop_cols if 'q9' in c.lower() and 'sofocos' in c.lower()]

shop['Fecha'] = pd.to_datetime(shop_raw[fecha_s[0]].apply(strip_shopify_quotes), errors='coerce', utc=True) if fecha_s else pd.NaT
shop['Symptoms_raw'] = shop_raw[symptoms_s[0]].apply(strip_shopify_quotes) if symptoms_s else ''
shop['Fase_raw'] = shop_raw[fase_s[0]].apply(strip_shopify_quotes) if fase_s else ''
shop['Canal_raw'] = shop_raw[canal_s[0]].apply(strip_shopify_quotes) if canal_s else ''
shop['Emocional_raw'] = shop_raw[emocional_s[0]].apply(strip_shopify_quotes) if emocional_s else ''
shop['Sexual_raw'] = shop_raw[sexual_s[0]].apply(strip_shopify_quotes) if sexual_s else ''
shop['Sofocos_cuando'] = shop_raw[sofocos_s[0]].apply(strip_shopify_quotes) if sofocos_s else ''

shop['Symptoms_list'] = shop['Symptoms_raw'].apply(parse_symptoms_slash)
for s in STANDARD_SYMPTOMS:
    shop[s] = shop['Symptoms_list'].apply(lambda x: 1 if s in x else 0)
shop['Num_Sintomas'] = shop[STANDARD_SYMPTOMS].sum(axis=1)

shop_emo = shop['Emocional_raw'].apply(parse_emotional_slash)
emo_df_s = pd.DataFrame(shop_emo.tolist())
for e in EMOTIONAL_COLS:
    shop[e] = emo_df_s[e] if e in emo_df_s.columns else 0

shop['Fase_Menstrual'] = shop['Fase_raw'].apply(normalize_fase)
shop['Canal'] = shop['Canal_raw'].apply(normalize_canal)

del shop_raw
print(f"  Shopify processed: {len(shop)} rows")

# --- TYPEFORM PRINCIPAL ---
print("Loading Typeform Principal data...")
tf_raw = pd.read_csv(TYPEFORM_PATH, low_memory=False)
print(f"  Typeform columns: {len(tf_raw.columns)}, rows: {len(tf_raw)}")

tf = pd.DataFrame()
tf['Fuente'] = 'Typeform'

tf_cols = tf_raw.columns.tolist()
# Date - Submit Date
fecha_tf = [c for c in tf_cols if 'submit date' in c.lower()]
tf['Fecha'] = pd.to_datetime(tf_raw[fecha_tf[0]], errors='coerce', utc=True) if fecha_tf else pd.NaT

# Phase
fase_tf = [c for c in tf_cols if 'menstruación' in c.lower()]
tf['Fase_raw'] = tf_raw[fase_tf[0]] if fase_tf else ''
tf['Fase_Menstrual'] = tf['Fase_raw'].apply(normalize_fase)

# Canal
canal_tf = [c for c in tf_cols if 'conocido' in c.lower()]
tf['Canal_raw'] = tf_raw[canal_tf[0]] if canal_tf else ''
tf['Canal'] = tf['Canal_raw'].apply(normalize_canal)

# Symptoms - individual boolean columns
symptom_tf_map = {
    'Sofocos': 'Sofocos',
    'Sequedad vaginal': 'Sequedad',
    'Sequedad': 'Sequedad',
    'Insomnio': 'Insomnio',
    'Baja libido': 'Baja libido',
    'Sudores nocturnos': 'Sudores nocturnos',
    'Aumento de peso': 'Aumento peso',
    'Fatiga': 'Fatiga',
}

for s in STANDARD_SYMPTOMS:
    tf[s] = 0

for tf_col_name, std_name in symptom_tf_map.items():
    matching = [c for c in tf_cols if c.strip() == tf_col_name]
    if matching:
        tf[std_name] = tf_raw[matching[0]].apply(lambda x: 1 if pd.notna(x) and str(x).strip() != '' and str(x).strip() != '0' else 0)
        # Ensure we use max in case Sequedad vaginal and Sequedad both map to Sequedad
        if std_name == 'Sequedad':
            existing = tf['Sequedad'].copy()
            new_val = tf_raw[matching[0]].apply(lambda x: 1 if pd.notna(x) and str(x).strip() != '' and str(x).strip() != '0' else 0)
            tf['Sequedad'] = np.maximum(existing, new_val)

tf['Num_Sintomas'] = tf[STANDARD_SYMPTOMS].sum(axis=1)

# Emotional - individual boolean columns
emo_tf_map = {
    'ANSIOSA': 'Ansiosa',
    'ESTRESADA': 'Estresada',
    'CAMBIOS DE HUMOR': 'Cambios_humor',
    'DESCENTRADA': 'Descentrada',
    'IRRITABLE': 'Irritable',
    'TRISTE': 'Triste',
}

for e in EMOTIONAL_COLS:
    tf[e] = 0

for tf_emo_name, std_emo in emo_tf_map.items():
    matching = [c for c in tf_cols if c.strip() == tf_emo_name]
    if matching:
        tf[std_emo] = tf_raw[matching[0]].apply(lambda x: 1 if pd.notna(x) and str(x).strip() != '' and str(x).strip() != '0' else 0)

# Age from birth date
birth_col = [c for c in tf_cols if 'naciste' in c.lower() or 'nacimiento' in c.lower()]
if birth_col:
    _bd = pd.to_datetime(tf_raw[birth_col[0]], format='mixed', errors='coerce')
    if _bd.dt.tz is not None:
        _bd = _bd.dt.tz_convert(None)
    tf['Birth_date'] = _bd
    tf['Edad'] = ((pd.Timestamp.now() - tf['Birth_date']).dt.days / 365.25).round(0)
    tf.loc[tf['Edad'] < 18, 'Edad'] = np.nan
    tf.loc[tf['Edad'] > 100, 'Edad'] = np.nan

tf['Symptoms_raw'] = ''
tf['Emocional_raw'] = ''
tf['Sexual_raw'] = ''
tf['Sofocos_cuando'] = ''

del tf_raw
print(f"  Typeform processed: {len(tf)} rows")

# --- TYPEFORM PHYSICAL ---
print("Loading Typeform Physical data...")
tf_phys = pd.read_csv(TYPEFORM_PHYS_PATH, low_memory=False)
print(f"  Typeform Physical columns: {len(tf_phys.columns)}, rows: {len(tf_phys)}")
# Keep this raw for sheet 9

# ========================================================
# 2. COMBINE DATA
# ========================================================
print("Combining data...")

keep_cols = ['Fuente', 'Fecha', 'Fase_Menstrual', 'Canal', 'Num_Sintomas'] + STANDARD_SYMPTOMS + EMOTIONAL_COLS

combined = pd.concat([
    woo[keep_cols],
    shop[keep_cols],
    tf[keep_cols]
], ignore_index=True)

# Convert dates to date only (no timezone)
combined['Fecha'] = combined['Fecha'].dt.tz_localize(None) if combined['Fecha'].dt.tz is not None else combined['Fecha']
combined['Fecha'] = combined['Fecha'].dt.date

print(f"Combined dataset: {len(combined)} rows")

# ========================================================
# 3. COMPUTE ANALYTICS
# ========================================================
print("Computing analytics...")

total_responses = len(combined)
total_woo = len(woo)
total_shop = len(shop)
total_tf = len(tf)
total_phys = len(tf_phys)

# Symptom ranking
symptom_counts = combined[STANDARD_SYMPTOMS].sum().sort_values(ascending=False)
symptom_pct = (symptom_counts / total_responses * 100).round(1)

# By source
symptom_by_source = {}
for src, df_src in [('WooCommerce', woo), ('Shopify', shop), ('Typeform', tf)]:
    counts = df_src[STANDARD_SYMPTOMS].sum().sort_values(ascending=False)
    pct = (counts / len(df_src) * 100).round(1)
    symptom_by_source[src] = pct

# Co-occurrence matrix
cooccurrence = combined[STANDARD_SYMPTOMS].T.dot(combined[STANDARD_SYMPTOMS])
np.fill_diagonal(cooccurrence.values, 0)

# Phase distribution
fase_dist = combined['Fase_Menstrual'].value_counts()
fase_pct = (fase_dist / total_responses * 100).round(1)

# Cross-tab: phase x symptoms
phase_symptom_cross = combined.groupby('Fase_Menstrual')[STANDARD_SYMPTOMS].mean().round(3) * 100

# Emotional analysis
emo_counts = combined[EMOTIONAL_COLS].sum().sort_values(ascending=False)
emo_pct = (emo_counts / total_responses * 100).round(1)

# Emotional by phase
emo_by_phase = combined.groupby('Fase_Menstrual')[EMOTIONAL_COLS].mean().round(3) * 100

# Segments
def segment(row):
    n = row['Num_Sintomas']
    emo_sum = sum(row[e] for e in EMOTIONAL_COLS)
    if n >= 5 and emo_sum >= 3:
        return 'Alto impacto'
    elif n >= 3:
        return 'Impacto moderado'
    elif n >= 1:
        return 'Impacto leve'
    else:
        return 'Sin síntomas reportados'

combined['Segmento'] = combined.apply(segment, axis=1)
segment_dist = combined['Segmento'].value_counts()
segment_pct = (segment_dist / total_responses * 100).round(1)

# Channel analysis
canal_dist = combined['Canal'].value_counts()
canal_pct = (canal_dist / total_responses * 100).round(1)

# Canal by source
canal_by_source = combined.groupby(['Canal', 'Fuente']).size().unstack(fill_value=0)

# Trends - quarterly
combined['Fecha_dt'] = pd.to_datetime(combined['Fecha'], errors='coerce')
combined['Quarter'] = combined['Fecha_dt'].dt.to_period('Q').astype(str)
quarterly = combined.groupby('Quarter').agg(
    Total=('Fuente', 'size'),
    Avg_Sintomas=('Num_Sintomas', 'mean'),
).round(2)
quarterly_source = combined.groupby(['Quarter', 'Fuente']).size().unstack(fill_value=0)

# Average symptoms
avg_symptoms = combined['Num_Sintomas'].mean()

# Top phase
top_fase = fase_dist.index[0] if len(fase_dist) > 0 else 'N/A'
top_canal = canal_dist.index[0] if len(canal_dist) > 0 else 'N/A'
top_symptom = symptom_counts.index[0] if len(symptom_counts) > 0 else 'N/A'

# ========================================================
# 4. BUILD EXCEL WORKBOOK
# ========================================================
print("Building Excel workbook...")

wb = Workbook()

def style_header_row(ws, row_num, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def style_data_rows(ws, start_row, end_row, max_col):
    for row in range(start_row, end_row + 1):
        fill = cream_fill if (row - start_row) % 2 == 0 else white_fill
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

def auto_width(ws, max_col, max_width=35):
    for col in range(1, max_col + 1):
        max_len = 0
        letter = get_column_letter(col)
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + 3, max_width)

def write_df_to_sheet(ws, df, start_row=1, start_col=1, header=True):
    """Write a dataframe to worksheet starting at given position."""
    rows = dataframe_to_rows(df, index=False, header=header)
    for r_idx, row in enumerate(rows, start_row):
        for c_idx, value in enumerate(row, start_col):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

def add_table(ws, start_row, end_row, start_col, end_col, name):
    """Add Excel table to range."""
    ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    safe_name = re.sub(r'[^A-Za-z0-9_]', '_', name)
    table = Table(displayName=safe_name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                          showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)

# ---- SHEET 1: RESUMEN ----
print("  Creating RESUMEN sheet...")
ws = wb.active
ws.title = "RESUMEN"
ws.sheet_properties.tabColor = DARK_BLUE

# Title
ws.merge_cells('A1:H1')
ws['A1'] = "DOMMA - Analisis Business Analytics Menopausia"
ws['A1'].font = title_font
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 40

ws.merge_cells('A2:H2')
ws['A2'] = f"Fecha de generacion: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Total respuestas: {total_responses:,}"
ws['A2'].font = Font(name='Calibri', color="666666", size=11)
ws['A2'].alignment = Alignment(horizontal='center')

# Key metrics row
metrics = [
    ("Total Respuestas", f"{total_responses:,}"),
    ("WooCommerce", f"{total_woo:,}"),
    ("Shopify", f"{total_shop:,}"),
    ("Typeform", f"{total_tf:,}"),
    ("Estudio Fisico", f"{total_phys:,}"),
    ("Promedio Sintomas", f"{avg_symptoms:.1f}"),
    ("Top Sintoma", top_symptom),
    ("Top Canal", top_canal),
]

row = 4
for col_idx, (label, value) in enumerate(metrics, 1):
    cell_val = ws.cell(row=row, column=col_idx, value=value)
    cell_val.font = Font(name='Calibri', bold=True, color=DARK_BLUE, size=18)
    cell_val.alignment = Alignment(horizontal='center', vertical='center')
    cell_val.fill = pink_fill
    cell_val.border = thin_border

    cell_lbl = ws.cell(row=row + 1, column=col_idx, value=label)
    cell_lbl.font = metric_label_font
    cell_lbl.alignment = Alignment(horizontal='center')

ws.row_dimensions[4].height = 40
ws.row_dimensions[5].height = 20

# Top 10 symptoms summary
row = 7
ws.cell(row=row, column=1, value="TOP SINTOMAS").font = subtitle_font
row = 8
ws.cell(row=row, column=1, value="Sintoma").font = header_font
ws.cell(row=row, column=1).fill = header_fill
ws.cell(row=row, column=2, value="Cantidad").font = header_font
ws.cell(row=row, column=2).fill = header_fill
ws.cell(row=row, column=3, value="%").font = header_font
ws.cell(row=row, column=3).fill = header_fill

for i, (symptom, count) in enumerate(symptom_counts.items()):
    r = row + 1 + i
    ws.cell(row=r, column=1, value=symptom)
    ws.cell(row=r, column=2, value=int(count))
    ws.cell(row=r, column=3, value=round(symptom_pct[symptom], 1))
    fill = cream_fill if i % 2 == 0 else white_fill
    for c in range(1, 4):
        ws.cell(row=r, column=c).fill = fill
        ws.cell(row=r, column=c).border = thin_border

# Phase distribution
phase_start_row = 7
ws.cell(row=phase_start_row, column=5, value="DISTRIBUCION POR FASE").font = subtitle_font
ws.cell(row=phase_start_row + 1, column=5, value="Fase").font = header_font
ws.cell(row=phase_start_row + 1, column=5).fill = header_fill
ws.cell(row=phase_start_row + 1, column=6, value="Cantidad").font = header_font
ws.cell(row=phase_start_row + 1, column=6).fill = header_fill
ws.cell(row=phase_start_row + 1, column=7, value="%").font = header_font
ws.cell(row=phase_start_row + 1, column=7).fill = header_fill

for i, (fase, count) in enumerate(fase_dist.items()):
    r = phase_start_row + 2 + i
    ws.cell(row=r, column=5, value=fase)
    ws.cell(row=r, column=6, value=int(count))
    ws.cell(row=r, column=7, value=round(fase_pct[fase], 1))
    fill = cream_fill if i % 2 == 0 else white_fill
    for c in range(5, 8):
        ws.cell(row=r, column=c).fill = fill
        ws.cell(row=r, column=c).border = thin_border

auto_width(ws, 8)

# ---- SHEET 2: DATOS_COMBINADOS ----
print("  Creating DATOS_COMBINADOS sheet...")
ws2 = wb.create_sheet("DATOS_COMBINADOS")
ws2.sheet_properties.tabColor = DARK_BLUE

output_cols = ['Fuente', 'Fecha', 'Fase_Menstrual', 'Canal', 'Num_Sintomas'] + STANDARD_SYMPTOMS + EMOTIONAL_COLS
combined_out = combined[output_cols].copy()

# Write headers
for col_idx, col_name in enumerate(output_cols, 1):
    ws2.cell(row=1, column=col_idx, value=col_name)
style_header_row(ws2, 1, len(output_cols))

# Write data in chunks to manage memory
print("  Writing combined data rows...")
chunk_size = 5000
for start in range(0, len(combined_out), chunk_size):
    chunk = combined_out.iloc[start:start + chunk_size]
    for r_idx, row_data in enumerate(chunk.itertuples(index=False), start + 2):
        for c_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=r_idx, column=c_idx)
            if isinstance(value, (pd.Timestamp, datetime)):
                cell.value = value
                cell.number_format = 'YYYY-MM-DD'
            elif isinstance(value, (np.integer, np.int64)):
                cell.value = int(value)
            elif isinstance(value, (np.floating, np.float64)):
                cell.value = float(value) if not np.isnan(value) else None
            else:
                cell.value = value

end_row = len(combined_out) + 1
# Add table
try:
    add_table(ws2, 1, end_row, 1, len(output_cols), "DatosCombinados")
except:
    pass

# Set column widths
for col_idx in range(1, len(output_cols) + 1):
    letter = get_column_letter(col_idx)
    ws2.column_dimensions[letter].width = max(len(output_cols[col_idx - 1]) + 2, 12)

# Freeze panes
ws2.freeze_panes = 'A2'
print(f"  Combined data written: {len(combined_out)} rows")

# ---- SHEET 3: SINTOMAS_RANKING ----
print("  Creating SINTOMAS_RANKING sheet...")
ws3 = wb.create_sheet("SINTOMAS_RANKING")
ws3.sheet_properties.tabColor = PINK

ws3.merge_cells('A1:F1')
ws3['A1'] = "Ranking de Sintomas"
ws3['A1'].font = title_font

# Main ranking
headers = ['Rank', 'Sintoma', 'Total', '%', 'WooCommerce %', 'Shopify %', 'Typeform %']
for col_idx, h in enumerate(headers, 1):
    ws3.cell(row=2, column=col_idx, value=h)
style_header_row(ws3, 2, len(headers))

for rank, (symptom, count) in enumerate(symptom_counts.items(), 1):
    r = 2 + rank
    ws3.cell(row=r, column=1, value=rank)
    ws3.cell(row=r, column=2, value=symptom)
    ws3.cell(row=r, column=3, value=int(count))
    ws3.cell(row=r, column=4, value=round(symptom_pct[symptom], 1))
    ws3.cell(row=r, column=5, value=round(symptom_by_source.get('WooCommerce', pd.Series()).get(symptom, 0), 1))
    ws3.cell(row=r, column=6, value=round(symptom_by_source.get('Shopify', pd.Series()).get(symptom, 0), 1))
    ws3.cell(row=r, column=7, value=round(symptom_by_source.get('Typeform', pd.Series()).get(symptom, 0), 1))

style_data_rows(ws3, 3, 2 + len(symptom_counts), len(headers))

# Add color scale for percentages
pct_range = f"D3:G{2 + len(symptom_counts)}"
ws3.conditional_formatting.add(pct_range, ColorScaleRule(
    start_type='min', start_color='FFFFFF',
    mid_type='percentile', mid_value=50, mid_color=PINK,
    end_type='max', end_color=DARK_BLUE
))

# Co-occurrence section
cooc_start = 2 + len(symptom_counts) + 3
ws3.cell(row=cooc_start, column=1, value="Matriz de Co-ocurrencia (top sintomas)").font = subtitle_font

top_symptoms = symptom_counts.head(8).index.tolist()
cooc_sub = cooccurrence.loc[top_symptoms, top_symptoms]

# Headers
ws3.cell(row=cooc_start + 1, column=1, value="")
for ci, s in enumerate(top_symptoms, 2):
    ws3.cell(row=cooc_start + 1, column=ci, value=s)
style_header_row(ws3, cooc_start + 1, len(top_symptoms) + 1)

for ri, s_row in enumerate(top_symptoms):
    r = cooc_start + 2 + ri
    ws3.cell(row=r, column=1, value=s_row).font = Font(bold=True)
    ws3.cell(row=r, column=1).fill = pink_fill
    for ci, s_col in enumerate(top_symptoms, 2):
        ws3.cell(row=r, column=ci, value=int(cooc_sub.loc[s_row, s_col]))

style_data_rows(ws3, cooc_start + 2, cooc_start + 1 + len(top_symptoms), len(top_symptoms) + 1)
auto_width(ws3, len(top_symptoms) + 1)

# ---- SHEET 4: CRUCES_FASE ----
print("  Creating CRUCES_FASE sheet...")
ws4 = wb.create_sheet("CRUCES_FASE")
ws4.sheet_properties.tabColor = PINK

ws4.merge_cells('A1:M1')
ws4['A1'] = "Cruces: Fase Menstrual x Sintomas (%)"
ws4['A1'].font = title_font

# Write cross-tab
cross = phase_symptom_cross.reset_index()
cross_headers = ['Fase_Menstrual'] + STANDARD_SYMPTOMS
for ci, h in enumerate(cross_headers, 1):
    ws4.cell(row=2, column=ci, value=h)
style_header_row(ws4, 2, len(cross_headers))

for ri, row_data in cross.iterrows():
    r = 3 + ri
    ws4.cell(row=r, column=1, value=row_data['Fase_Menstrual'])
    ws4.cell(row=r, column=1).font = Font(bold=True)
    for ci, s in enumerate(STANDARD_SYMPTOMS, 2):
        val = round(row_data[s], 1) if s in row_data.index else 0
        ws4.cell(row=r, column=ci, value=val)

end_r = 2 + len(cross)
style_data_rows(ws4, 3, end_r, len(cross_headers))

# Color scale
data_range = f"B3:{get_column_letter(len(cross_headers))}{end_r}"
ws4.conditional_formatting.add(data_range, ColorScaleRule(
    start_type='min', start_color='FFFFFF',
    mid_type='percentile', mid_value=50, mid_color=PINK,
    end_type='max', end_color=DARK_BLUE
))

auto_width(ws4, len(cross_headers))
ws4.freeze_panes = 'B3'

# ---- SHEET 5: EMOCIONAL ----
print("  Creating EMOCIONAL sheet...")
ws5 = wb.create_sheet("EMOCIONAL")
ws5.sheet_properties.tabColor = PINK

ws5.merge_cells('A1:G1')
ws5['A1'] = "Analisis Emocional"
ws5['A1'].font = title_font

# Overall emotional counts
headers_emo = ['Estado Emocional', 'Total', '%']
for ci, h in enumerate(headers_emo, 1):
    ws5.cell(row=2, column=ci, value=h)
style_header_row(ws5, 2, len(headers_emo))

for i, (emo, count) in enumerate(emo_counts.items()):
    r = 3 + i
    ws5.cell(row=r, column=1, value=emo)
    ws5.cell(row=r, column=2, value=int(count))
    ws5.cell(row=r, column=3, value=round(emo_pct[emo], 1))

style_data_rows(ws5, 3, 2 + len(emo_counts), 3)

# Emotional by phase
emo_phase_start = 2 + len(emo_counts) + 3
ws5.cell(row=emo_phase_start, column=1, value="Estado Emocional por Fase (%)").font = subtitle_font

emo_phase = emo_by_phase.reset_index()
emo_phase_headers = ['Fase_Menstrual'] + EMOTIONAL_COLS
for ci, h in enumerate(emo_phase_headers, 1):
    ws5.cell(row=emo_phase_start + 1, column=ci, value=h)
style_header_row(ws5, emo_phase_start + 1, len(emo_phase_headers))

for ri, row_data in emo_phase.iterrows():
    r = emo_phase_start + 2 + ri
    ws5.cell(row=r, column=1, value=row_data['Fase_Menstrual'])
    ws5.cell(row=r, column=1).font = Font(bold=True)
    for ci, e in enumerate(EMOTIONAL_COLS, 2):
        ws5.cell(row=r, column=ci, value=round(row_data[e], 1))

end_emo = emo_phase_start + 1 + len(emo_phase)
style_data_rows(ws5, emo_phase_start + 2, end_emo, len(emo_phase_headers))

# Color scale
emo_range = f"B{emo_phase_start + 2}:{get_column_letter(len(emo_phase_headers))}{end_emo}"
ws5.conditional_formatting.add(emo_range, ColorScaleRule(
    start_type='min', start_color='FFFFFF',
    mid_type='percentile', mid_value=50, mid_color=PINK,
    end_type='max', end_color=DARK_BLUE
))

auto_width(ws5, len(emo_phase_headers))

# ---- SHEET 6: SEGMENTOS ----
print("  Creating SEGMENTOS sheet...")
ws6 = wb.create_sheet("SEGMENTOS")
ws6.sheet_properties.tabColor = PINK

ws6.merge_cells('A1:E1')
ws6['A1'] = "Segmentacion por Impacto"
ws6['A1'].font = title_font

seg_headers = ['Segmento', 'Total', '%', 'Promedio Sintomas']
for ci, h in enumerate(seg_headers, 1):
    ws6.cell(row=2, column=ci, value=h)
style_header_row(ws6, 2, len(seg_headers))

seg_stats = combined.groupby('Segmento').agg(
    Total=('Fuente', 'size'),
    Avg_Sintomas=('Num_Sintomas', 'mean')
).sort_values('Total', ascending=False)

for i, (seg, row_data) in enumerate(seg_stats.iterrows()):
    r = 3 + i
    ws6.cell(row=r, column=1, value=seg)
    ws6.cell(row=r, column=2, value=int(row_data['Total']))
    ws6.cell(row=r, column=3, value=round(row_data['Total'] / total_responses * 100, 1))
    ws6.cell(row=r, column=4, value=round(row_data['Avg_Sintomas'], 1))

style_data_rows(ws6, 3, 2 + len(seg_stats), len(seg_headers))

# Segment by source
seg_src_start = 2 + len(seg_stats) + 3
ws6.cell(row=seg_src_start, column=1, value="Segmentos por Fuente").font = subtitle_font

seg_src = combined.groupby(['Segmento', 'Fuente']).size().unstack(fill_value=0)
seg_src_headers = ['Segmento'] + seg_src.columns.tolist()
for ci, h in enumerate(seg_src_headers, 1):
    ws6.cell(row=seg_src_start + 1, column=ci, value=h)
style_header_row(ws6, seg_src_start + 1, len(seg_src_headers))

for i, (seg, row_data) in enumerate(seg_src.iterrows()):
    r = seg_src_start + 2 + i
    ws6.cell(row=r, column=1, value=seg)
    for ci, src in enumerate(seg_src.columns, 2):
        ws6.cell(row=r, column=ci, value=int(row_data[src]))

style_data_rows(ws6, seg_src_start + 2, seg_src_start + 1 + len(seg_src), len(seg_src_headers))

# Segment by phase
seg_fase_start = seg_src_start + len(seg_src) + 3
ws6.cell(row=seg_fase_start, column=1, value="Segmentos por Fase Menstrual").font = subtitle_font

seg_fase = combined.groupby(['Segmento', 'Fase_Menstrual']).size().unstack(fill_value=0)
seg_fase_headers = ['Segmento'] + seg_fase.columns.tolist()
for ci, h in enumerate(seg_fase_headers, 1):
    ws6.cell(row=seg_fase_start + 1, column=ci, value=h)
style_header_row(ws6, seg_fase_start + 1, len(seg_fase_headers))

for i, (seg, row_data) in enumerate(seg_fase.iterrows()):
    r = seg_fase_start + 2 + i
    ws6.cell(row=r, column=1, value=seg)
    for ci, fase in enumerate(seg_fase.columns, 2):
        ws6.cell(row=r, column=ci, value=int(row_data[fase]))

style_data_rows(ws6, seg_fase_start + 2, seg_fase_start + 1 + len(seg_fase), len(seg_fase_headers))
auto_width(ws6, max(len(seg_headers), len(seg_src_headers), len(seg_fase_headers)))

# ---- SHEET 7: CANALES ----
print("  Creating CANALES sheet...")
ws7 = wb.create_sheet("CANALES")
ws7.sheet_properties.tabColor = PINK

ws7.merge_cells('A1:F1')
ws7['A1'] = "Analisis de Canales de Adquisicion"
ws7['A1'].font = title_font

canal_headers = ['Canal', 'Total', '%']
for ci, h in enumerate(canal_headers, 1):
    ws7.cell(row=2, column=ci, value=h)
style_header_row(ws7, 2, len(canal_headers))

for i, (canal, count) in enumerate(canal_dist.items()):
    r = 3 + i
    ws7.cell(row=r, column=1, value=canal)
    ws7.cell(row=r, column=2, value=int(count))
    ws7.cell(row=r, column=3, value=round(canal_pct[canal], 1))

style_data_rows(ws7, 3, 2 + len(canal_dist), len(canal_headers))

# Canal by source
canal_src_start = 2 + len(canal_dist) + 3
ws7.cell(row=canal_src_start, column=1, value="Canales por Fuente").font = subtitle_font

canal_src_headers = ['Canal'] + canal_by_source.columns.tolist()
for ci, h in enumerate(canal_src_headers, 1):
    ws7.cell(row=canal_src_start + 1, column=ci, value=h)
style_header_row(ws7, canal_src_start + 1, len(canal_src_headers))

for i, (canal, row_data) in enumerate(canal_by_source.iterrows()):
    r = canal_src_start + 2 + i
    ws7.cell(row=r, column=1, value=canal)
    for ci, src in enumerate(canal_by_source.columns, 2):
        ws7.cell(row=r, column=ci, value=int(row_data[src]))

style_data_rows(ws7, canal_src_start + 2, canal_src_start + 1 + len(canal_by_source), len(canal_src_headers))

# Canal x avg symptoms
canal_symp_start = canal_src_start + len(canal_by_source) + 3
ws7.cell(row=canal_symp_start, column=1, value="Promedio Sintomas por Canal").font = subtitle_font
canal_avg = combined.groupby('Canal')['Num_Sintomas'].mean().sort_values(ascending=False).round(2)
cs_headers = ['Canal', 'Promedio Sintomas']
for ci, h in enumerate(cs_headers, 1):
    ws7.cell(row=canal_symp_start + 1, column=ci, value=h)
style_header_row(ws7, canal_symp_start + 1, len(cs_headers))

for i, (canal, avg) in enumerate(canal_avg.items()):
    r = canal_symp_start + 2 + i
    ws7.cell(row=r, column=1, value=canal)
    ws7.cell(row=r, column=2, value=round(avg, 2))

style_data_rows(ws7, canal_symp_start + 2, canal_symp_start + 1 + len(canal_avg), len(cs_headers))
auto_width(ws7, max(len(canal_headers), len(canal_src_headers)))

# ---- SHEET 8: TENDENCIAS ----
print("  Creating TENDENCIAS sheet...")
ws8 = wb.create_sheet("TENDENCIAS")
ws8.sheet_properties.tabColor = PINK

ws8.merge_cells('A1:F1')
ws8['A1'] = "Tendencias Trimestrales"
ws8['A1'].font = title_font

# Volume trends
trend_headers = ['Trimestre', 'Total', 'Promedio Sintomas']
for ci, h in enumerate(trend_headers, 1):
    ws8.cell(row=2, column=ci, value=h)
style_header_row(ws8, 2, len(trend_headers))

quarterly_sorted = quarterly.sort_index()
for i, (q, row_data) in enumerate(quarterly_sorted.iterrows()):
    r = 3 + i
    ws8.cell(row=r, column=1, value=str(q))
    ws8.cell(row=r, column=2, value=int(row_data['Total']))
    ws8.cell(row=r, column=3, value=round(row_data['Avg_Sintomas'], 2))

style_data_rows(ws8, 3, 2 + len(quarterly_sorted), len(trend_headers))

# Volume by source over time
src_trend_start = 2 + len(quarterly_sorted) + 3
ws8.cell(row=src_trend_start, column=1, value="Volumen por Fuente y Trimestre").font = subtitle_font

quarterly_source_sorted = quarterly_source.sort_index()
qs_headers = ['Trimestre'] + quarterly_source_sorted.columns.tolist()
for ci, h in enumerate(qs_headers, 1):
    ws8.cell(row=src_trend_start + 1, column=ci, value=h)
style_header_row(ws8, src_trend_start + 1, len(qs_headers))

for i, (q, row_data) in enumerate(quarterly_source_sorted.iterrows()):
    r = src_trend_start + 2 + i
    ws8.cell(row=r, column=1, value=str(q))
    for ci, src in enumerate(quarterly_source_sorted.columns, 2):
        ws8.cell(row=r, column=ci, value=int(row_data[src]))

style_data_rows(ws8, src_trend_start + 2, src_trend_start + 1 + len(quarterly_source_sorted), len(qs_headers))
auto_width(ws8, max(len(trend_headers), len(qs_headers)))

# ---- SHEET 9: ESTUDIO_FISICO ----
print("  Creating ESTUDIO_FISICO sheet...")
ws9 = wb.create_sheet("ESTUDIO_FISICO")
ws9.sheet_properties.tabColor = DARK_BLUE

ws9.merge_cells(f'A1:{get_column_letter(len(tf_phys.columns))}1')
ws9['A1'] = "Estudio Fisico - Typeform (729 respuestas)"
ws9['A1'].font = title_font

# Write all physical data
phys_cols = tf_phys.columns.tolist()
for ci, h in enumerate(phys_cols, 1):
    ws9.cell(row=2, column=ci, value=h)
style_header_row(ws9, 2, len(phys_cols))

for ri, row_data in tf_phys.iterrows():
    r = 3 + ri
    for ci, col in enumerate(phys_cols, 1):
        val = row_data[col]
        if pd.isna(val):
            ws9.cell(row=r, column=ci, value="")
        else:
            ws9.cell(row=r, column=ci, value=str(val)[:200])

style_data_rows(ws9, 3, 2 + len(tf_phys), len(phys_cols))
auto_width(ws9, len(phys_cols), max_width=30)
ws9.freeze_panes = 'A3'

# ---- SHEET 10: PARA_FARMA ----
print("  Creating PARA_FARMA sheet...")
ws10 = wb.create_sheet("PARA_FARMA")
ws10.sheet_properties.tabColor = DARK_BLUE

ws10.merge_cells('A1:F1')
ws10['A1'] = "Datos Clave para Presentacion Farma"
ws10['A1'].font = title_font

ws10.merge_cells('A2:F2')
ws10['A2'] = f"Base: {total_responses:,} mujeres | Periodo: 2024-2026"
ws10['A2'].font = Font(name='Calibri', color="666666", size=11, italic=True)

# Key data points
r = 4
ws10.cell(row=r, column=1, value="DATO CLAVE").font = subtitle_font
ws10.cell(row=r, column=3, value="VALOR").font = subtitle_font
ws10.cell(row=r, column=5, value="CONTEXTO").font = subtitle_font

farma_data = [
    ("Tamano de muestra", f"{total_responses:,} mujeres", "Datos recopilados via quiz online (WooCommerce, Shopify, Typeform)"),
    ("Sintoma mas prevalente", f"{top_symptom} ({symptom_pct.iloc[0]:.0f}%)", f"Sobre {total_responses:,} respuestas totales"),
    ("Promedio sintomas por mujer", f"{avg_symptoms:.1f} sintomas", "Indica carga sintomatica significativa"),
    ("% con 3+ sintomas", f"{(combined['Num_Sintomas'] >= 3).mean() * 100:.0f}%", "Mayoria con multiples sintomas simultaneos"),
    ("% con 5+ sintomas (alto impacto)", f"{(combined['Num_Sintomas'] >= 5).mean() * 100:.0f}%", "Segmento de alta necesidad"),
    ("Fase mas representada", f"{top_fase} ({fase_pct.iloc[0]:.0f}%)", "Etapa hormonal predominante en la muestra"),
    ("Canal principal de captacion", f"{top_canal} ({canal_pct.iloc[0]:.0f}%)", "Como llegan las mujeres al quiz"),
    ("Top emocion reportada", f"{emo_counts.index[0]} ({emo_pct.iloc[0]:.0f}%)", "Estado emocional mas frecuente"),
    ("Estudio fisico adicional", f"{total_phys} respuestas", "Encuesta complementaria sobre salud fisica"),
]

for i, (dato, valor, contexto) in enumerate(farma_data):
    r = 5 + i
    ws10.cell(row=r, column=1, value=dato).font = Font(bold=True, name='Calibri', size=11)
    ws10.cell(row=r, column=3, value=valor).font = Font(name='Calibri', size=11, color=DARK_BLUE, bold=True)
    ws10.cell(row=r, column=5, value=contexto).font = Font(name='Calibri', size=10, color="666666")
    fill = cream_fill if i % 2 == 0 else white_fill
    for c in range(1, 7):
        ws10.cell(row=r, column=c).fill = fill
        ws10.cell(row=r, column=c).border = thin_border

# Top 5 symptoms for pharma
farma_symp_start = 5 + len(farma_data) + 2
ws10.cell(row=farma_symp_start, column=1, value="TOP 5 SINTOMAS - RESUMEN FARMA").font = subtitle_font

farma_symp_headers = ['#', 'Sintoma', 'Prevalencia %', 'N']
for ci, h in enumerate(farma_symp_headers, 1):
    ws10.cell(row=farma_symp_start + 1, column=ci, value=h)
style_header_row(ws10, farma_symp_start + 1, len(farma_symp_headers))

for i, (symptom, count) in enumerate(symptom_counts.head(5).items()):
    r = farma_symp_start + 2 + i
    ws10.cell(row=r, column=1, value=i + 1)
    ws10.cell(row=r, column=2, value=symptom)
    ws10.cell(row=r, column=3, value=round(symptom_pct[symptom], 1))
    ws10.cell(row=r, column=4, value=int(count))

style_data_rows(ws10, farma_symp_start + 2, farma_symp_start + 6, len(farma_symp_headers))

# Top co-occurring pairs for pharma
pair_start = farma_symp_start + 9
ws10.cell(row=pair_start, column=1, value="TOP COMBINACIONES DE SINTOMAS").font = subtitle_font

# Get top pairs
pairs = []
for i, s1 in enumerate(STANDARD_SYMPTOMS):
    for j, s2 in enumerate(STANDARD_SYMPTOMS):
        if i < j and cooccurrence.loc[s1, s2] > 0:
            pairs.append((s1, s2, int(cooccurrence.loc[s1, s2])))
pairs.sort(key=lambda x: x[2], reverse=True)

pair_headers = ['Sintoma 1', 'Sintoma 2', 'Co-ocurrencia (N)']
for ci, h in enumerate(pair_headers, 1):
    ws10.cell(row=pair_start + 1, column=ci, value=h)
style_header_row(ws10, pair_start + 1, len(pair_headers))

for i, (s1, s2, n) in enumerate(pairs[:8]):
    r = pair_start + 2 + i
    ws10.cell(row=r, column=1, value=s1)
    ws10.cell(row=r, column=2, value=s2)
    ws10.cell(row=r, column=3, value=n)

style_data_rows(ws10, pair_start + 2, pair_start + 9, len(pair_headers))
auto_width(ws10, 6, max_width=40)

# ========================================================
# 5. SAVE
# ========================================================
print("Saving workbook...")
wb.save(OUTPUT_PATH)
print(f"Workbook saved to: {OUTPUT_PATH}")

# Verify
import os
size_mb = os.path.getsize(OUTPUT_PATH) / (1024 * 1024)
print(f"File size: {size_mb:.1f} MB")
print("Done!")
